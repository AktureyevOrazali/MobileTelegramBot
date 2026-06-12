"""FastAPI routes serving chat data for the mobile client."""

from __future__ import annotations



import logging

import base64
import hashlib
import html

import hmac

import io

import json

import os

import re
import tempfile
from pathlib import Path

from datetime import date, datetime, timezone

from typing import Any, Dict, List, Optional, Literal

import asyncio

from collections import defaultdict
from urllib.parse import urlsplit, urlunsplit


from fastapi import APIRouter, BackgroundTasks, Depends, FastAPI, File, Form, Header, HTTPException, Query, Request, UploadFile

from fastapi.middleware.cors import CORSMiddleware

from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse

from fastapi.security import HTTPBearer

from pydantic import AliasChoices, BaseModel, EmailStr, Field, validator



from . import database
from . import employee_client_assessments
from . import survey_service

from .ai_manager import ai_manager

from .media import MediaValidationError, media_service, verify_media_access

from .telegram_bot import bot, enable_ai_session, send_ai_csat_request, send_csat_request

from . import contract_checker



# ------------------------ Temporary import for Aziret's employee cabinet ------------------

from .kabinet_sotrudnika_by_Aziret import kabinet_backend



API_TOKEN = os.getenv("MOBILE_API_TOKEN")

ONEC_INTEGRATION_TOKEN = os.getenv("ONEC_INTEGRATION_TOKEN")
MEDIA_PUBLIC_BASE_URL = (os.getenv("MEDIA_PUBLIC_BASE_URL", "").strip().rstrip("/") or os.getenv("PUBLIC_BASE_URL", "").strip().rstrip("/"))


ONEC_CHAT_ID_OFFSET = 9_000_000_000_000

ONEC_CHAT_ID_SPACE = 1_000_000_000_000



# Опциональный общий секрет для подписи HMAC нагрузки, которую 1С забирает из outbox

ONEC_SHARED_SECRET = os.getenv("ONEC_SHARED_SECRET", "")


ONEC_OPERATOR_COMMAND_VALUES = {
    "/operator",
    "operator",
    "позвать оператора",
    "позовите оператора",
    "позовите, пожалуйста, оператора.",
    "нужен оператор",
    "оператор",
    "операторды шақыру",
    "оператор керек",
}

ONEC_LANGUAGE_COMMANDS = {
    "/lang_ru": ("ru", "Язык диалога: русский."),
    "/lang_kk": ("kk", "Диалог тілі: қазақша."),
}


def _onec_quick_reply(
    reply_id: str,
    label: str,
    value: str,
    *,
    kind: str = "command",
) -> dict[str, str]:
    return {"id": reply_id, "label": label, "value": value, "type": kind}


def _onec_language_quick_replies() -> list[dict[str, str]]:
    return [
        _onec_quick_reply("lang_ru", "Русский", "/lang_ru", kind="language"),
        _onec_quick_reply("lang_kk", "Қазақша", "/lang_kk", kind="language"),
    ]


def _onec_default_quick_replies() -> list[dict[str, str]]:
    return [
        _onec_quick_reply("status", "Статус заявки", "Проверьте, пожалуйста, статус моей заявки."),
        _onec_quick_reply("docs", "Документы", "Какие документы нужны для оформления?"),
        _onec_quick_reply("operator", "Позвать оператора", "/operator", kind="operator_request"),
    ]


def _onec_rating_quick_replies(target: str) -> list[dict[str, str]]:
    return [
        _onec_quick_reply(
            f"rate_{target}_{rating}",
            str(rating),
            str(rating),
            kind=f"{target}_rating",
        )
        for rating in range(1, 6)
    ]


def _is_onec_operator_request(normalized_text: str) -> bool:
    text = (normalized_text or "").strip().lower()
    return text in ONEC_OPERATOR_COMMAND_VALUES


def _normalize_onec_language_command(normalized_text: str) -> tuple[str, str] | None:
    return ONEC_LANGUAGE_COMMANDS.get((normalized_text or "").strip().lower())


def _build_onec_contract_status_text(*, has_contract: bool, year: int | str) -> str:
    if has_contract:
        return f"Действующий договор на {year} год найден. Можете продолжить обращение."
    return (
        f"Не найден действующий договор на {year} год для этой организации.\n"
        "Для продолжения обслуживания обратитесь в офис для оформления договора."
    )



CORS_ORIGINS = [o.strip() for o in os.getenv("CORS_ORIGINS", "").split(",") if o.strip()]
if not CORS_ORIGINS:
    CORS_ORIGINS = [
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:4173",
        "http://127.0.0.1:4173",
    ]

CORS_ORIGIN_REGEX = (
    os.getenv("CORS_ORIGIN_REGEX", "").strip()
    or r"^https?://(localhost|127\.0\.0\.1|0\.0\.0\.0|10(?:\.\d{1,3}){3}|192\.168(?:\.\d{1,3}){2}|172\.(?:1[6-9]|2\d|3[0-1])(?:\.\d{1,3}){2})(?::\d+)?$"
)

app = FastAPI(title="MobileBot Companion API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_origin_regex=CORS_ORIGIN_REGEX,
    # The frontend sends custom auth headers, so credentials must stay enabled.
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)



logger = logging.getLogger(__name__)



router = APIRouter(prefix="/api")

security = HTTPBearer()



# ── SSE Event Bus ──

class EventBus:

    def __init__(self):

        # Maps user_id -> list of queues

        self.connections: dict[int, list[asyncio.Queue]] = defaultdict(list)

        self.loop: asyncio.AbstractEventLoop | None = None



    async def publish(self, user_id: int, event_type: str, data: dict):

        if user_id in self.connections:

            message = {"type": event_type, "data": data}

            for q in self.connections[user_id]:

                await q.put(message)



    async def publish_all(self, event_type: str, data: dict):

        message = {"type": event_type, "data": data}

        for qs in self.connections.values():

            for q in qs:

                await q.put(message)



event_bus = EventBus()


@app.on_event("startup")
async def _register_event_bus_loop() -> None:
    event_bus.loop = asyncio.get_running_loop()
    logger.info("SSE event bus loop registered")


@app.on_event("shutdown")
async def _clear_event_bus_loop() -> None:
    event_bus.loop = None
    event_bus.connections.clear()
    logger.info("SSE event bus loop cleared")



ROLE_LABELS: Dict[str, str] = {

    database.ROLE_ADMIN: "\u0410\u0434\u043c\u0438\u043d\u0438\u0441\u0442\u0440\u0430\u0442\u043e\u0440",

    database.ROLE_MODERATOR: "\u041c\u043e\u0434\u0435\u0440\u0430\u0442\u043e\u0440",

    database.ROLE_OPERATOR: "\u041e\u043f\u0435\u0440\u0430\u0442\u043e\u0440",

    database.ROLE_HR: "\u041a\u0430\u0434\u0440\u043e\u0432\u0438\u043a",

}





def _sign_payload(payload: dict) -> str:

    """HMAC-SHA256 signature for a compact JSON payload. Returns an empty string when the shared secret is missing."""

    if not ONEC_SHARED_SECRET:

        return ""

    body = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))

    return hmac.new(ONEC_SHARED_SECRET.encode("utf-8"), body.encode("utf-8"), hashlib.sha256).hexdigest()



def _get_base64_content(media_id: int) -> str | None:
    """Read media file, optimize image, and return base64 content."""
    try:
        media = media_service.get_media_descriptor(media_id)
        if not media or media.kind != "image":
            return None
        local_path = media_service.get_local_path(media)
        if not local_path or not local_path.exists():
            return None

        # Try to use Pillow for optimization, fallback to raw if not available
        try:
            from PIL import Image
        except ImportError:
            with open(local_path, "rb") as f:
                return base64.b64encode(f.read()).decode("utf-8")

        with Image.open(local_path) as img:
            # Convert to RGB (standard for JPEG)
            if img.mode != "RGB":
                img = img.convert("RGB")

            # Resize if too large - 800px is enough for 1C and much faster to encode/decode
            if max(img.width or 0, img.height or 0) > 800:
                img.thumbnail((800, 800), Image.Resampling.BILINEAR)

            buffer = io.BytesIO()
            img.save(buffer, format="JPEG", quality=80, optimize=True)
            return base64.b64encode(buffer.getvalue()).decode("utf-8")
    except Exception as exc:
        logger.error("Failed to generate base64 for media %s: %s", media_id, exc)
        return None





def _enqueue_onec_outgoing_message(

    *,

    message_id: int,

    chat_id: int,

    dialog_id: int | None,

    external_chat_id: str,

    bin_value: str | None,

    text: str,

    author: str | None,

    section: str | None,

    attachments: List[dict] | None = None,

    direction: str = "outgoing",
    quick_replies: List[dict] | None = None,

) -> None:

    payload = {

        "external_chat_id": external_chat_id,

        "chat_id": chat_id,

        "dialog_id": dialog_id,

        "text": text,

        "author": author,

        "created_at": datetime.now(timezone.utc).isoformat(timespec="seconds") + "Z",

        "bin": bin_value,

        "section": section,

        "direction": direction,

        "attachments": attachments or [],

    }
    if quick_replies:
        payload["quick_replies"] = quick_replies

    signature = _sign_payload(payload)

    if signature:

        payload["signature"] = signature

    try:

        database.outbox_enqueue_onec(

            message_id=message_id,

            chat_id=chat_id,

            external_chat_id=external_chat_id,

            bin_value=bin_value,

            payload=payload,

        )

    except Exception as exc:  # pragma: no cover - best-effort side effect

        logger.exception("Failed to enqueue 1C outbox message: %s", exc)





class RegisterRequest(BaseModel):

    name: str = Field(min_length=2, max_length=100)

    email: EmailStr

    password: str = Field(min_length=5, max_length=100)





class LoginRequest(BaseModel):

    identifier: str = Field(min_length=1, max_length=150)

    password: str





class AttachmentResponse(BaseModel):

    id: int

    media_id: int

    kind: Literal["image", "video"]

    url: str

    preview_url: str | None = None

    mime_type: str

    size_bytes: int

    original_name: str

    width: int | None = None

    height: int | None = None

    duration_sec: float | None = None

    caption: str | None = None

    base64_content: str | None = None





class UploadResponse(BaseModel):

    status: str = Field(default="ok")

    media_id: int

    kind: Literal["image", "video"]

    url: str

    preview_url: str | None = None

    mime_type: str

    size_bytes: int

    original_name: str

    width: int | None = None

    height: int | None = None

    duration_sec: float | None = None

    base64_content: str | None = None





class ReplyRequest(BaseModel):

    chat_id: int

    text: str = ""

    dialog_id: int | None = None

    attachment_ids: List[int] = Field(default_factory=list)



    @validator("attachment_ids", pre=True, always=True)

    def _normalize_attachment_ids(cls, value):

        if value in (None, ""):

            return []

        if not isinstance(value, list):

            raise ValueError("attachment_ids must be an array")

        normalized: List[int] = []

        seen: set[int] = set()

        for item in value:

            parsed = int(item)

            if parsed <= 0 or parsed in seen:

                continue

            normalized.append(parsed)

            seen.add(parsed)

        return normalized



    @validator("text", pre=True, always=True)

    def _normalize_text(cls, value):

        return "" if value is None else str(value)



    @validator("dialog_id", pre=True, always=True)

    def _normalize_dialog_id(cls, value):

        return None if value in (None, "") else int(value)



    @validator("chat_id", pre=True, always=True)

    def _normalize_chat_id(cls, value):

        return int(value)



    @validator("attachment_ids")

    def _require_text_or_attachment(cls, value, values):

        if not (values.get("text") or "").strip() and not value:

            raise ValueError("Message must contain text or attachments")

        return value





class OneCIncomingMessageRequest(BaseModel):

    external_chat_id: str = Field(min_length=1, max_length=128)

    text: str = ""

    bin: str = Field(min_length=3, max_length=32)

    author: str | None = Field(default=None, max_length=150)

    title: str | None = Field(default=None, max_length=150)

    section: str | None = Field(default=None, max_length=50)

    chat_id: int | None = None

    attachment_ids: List[int] = Field(default_factory=list)



    @validator("attachment_ids", pre=True, always=True)

    def _normalize_attachment_ids(cls, value):

        if value in (None, ""):

            return []

        if not isinstance(value, list):

            raise ValueError("attachment_ids must be an array")

        normalized: List[int] = []

        seen: set[int] = set()

        for item in value:

            parsed = int(item)

            if parsed <= 0 or parsed in seen:

                continue

            normalized.append(parsed)

            seen.add(parsed)

        return normalized



    @validator("text", pre=True, always=True)

    def _normalize_text(cls, value):

        return "" if value is None else str(value)



    @validator("chat_id", pre=True, always=True)

    def _normalize_chat_id(cls, value):

        return None if value in (None, "") else int(value)



    @validator("attachment_ids")

    def _require_payload(cls, value, values):

        if not (values.get("text") or "").strip() and not value:

            raise ValueError("Message must contain text or attachments")

        return value





class OneCMessageEntry(BaseModel):

    message_id: int | None = None

    chat_id: int

    dialog_id: int | None = None

    direction: str

    text: str

    author: str | None = None

    created_at: str

    section: str | None = None

    attachments: List[AttachmentResponse] = Field(default_factory=list)

    quick_replies: List[dict] = Field(default_factory=list)





class OneCMessagesResponse(BaseModel):

    status: str = Field(default="ok")

    external_chat_id: str

    chat_id: int

    dialog_id: int | None = None

    messages: List[OneCMessageEntry] = Field(default_factory=list)





class BinAssignmentResponse(BaseModel):

    bin: str

    assigned_at: str

    expires_at: str | None = None

    assigned_by: int | None = None





class UserResponse(BaseModel):

    id: int

    email: EmailStr

    login: str

    name: str

    created_at: str

    job_title: str = ""

    organization: str = "ТОО Азия-Сервис"

    phone: str = ""

    bio: str = ""

    role: str

    is_approved: bool = True

    sections: List[str] = Field(default_factory=list)

    bins: List[BinAssignmentResponse] = Field(default_factory=list)

    favorite_dialog_ids: List[int] = []





class AuthResponse(BaseModel):

    token: str

    user: UserResponse





class RegisterResponse(BaseModel):

    status: str = "pending"

    message: str = "\u0420\u0435\u0433\u0438\u0441\u0442\u0440\u0430\u0446\u0438\u044f \u043e\u0442\u043f\u0440\u0430\u0432\u043b\u0435\u043d\u0430. \u041e\u0436\u0438\u0434\u0430\u0439\u0442\u0435 \u043f\u043e\u0434\u0442\u0432\u0435\u0440\u0436\u0434\u0435\u043d\u0438\u044f \u043c\u043e\u0434\u0435\u0440\u0430\u0442\u043e\u0440\u0430."





class PendingUserResponse(BaseModel):

    id: int

    email: EmailStr

    name: str

    created_at: str





class BinAssignmentRequest(BaseModel):

    bin: str = Field(min_length=1, max_length=100)

    expires_at: datetime | None = Field(

        default=None,

        validation_alias=AliasChoices("expires_at", "expiresAt"),

        serialization_alias="expires_at",

    )





class RoleUpdateRequest(BaseModel):

    role: str = Field(min_length=3, max_length=20)





class SectionsUpdateRequest(BaseModel):

    sections: List[str] = Field(default_factory=list)





class PasswordChangeRequest(BaseModel):

    current_password: str = Field(min_length=5, max_length=100)

    new_password: str = Field(min_length=5, max_length=100)





class PasswordResetRequest(BaseModel):

    new_password: str = Field(min_length=5, max_length=100)





class BinsUpdateRequest(BaseModel):

    bins: List[BinAssignmentRequest] = Field(default_factory=list)





class UnassignedBinResponse(BaseModel):

    bin: str

    open_dialogs: int

    has_contract: bool = True





class OrganizationWithoutContractResponse(BaseModel):

    customer_bin: str

    customer_legal_address: str | None = None

    customer_bank_name_ru: str | None = None
    customer_name_ru: str | None = None

    created_at: str





class BinDetailedResponse(BaseModel):

    bin: str

    has_contract: bool

    customer_legal_address: str | None = None

    customer_bank_name_ru: str | None = None
    customer_name_ru: str | None = None



class MessageResponse(BaseModel):

    id: int

    message_id: int | None = None

    chat_id: int

    direction: str

    text: str

    author: str | None

    created_at: str

    section: str | None = None

    section_title: str | None = None

    dialog_id: int | None = None

    attachments: List[AttachmentResponse] = Field(default_factory=list)





class ChatResponse(BaseModel):

    chat_id: int

    dialog_id: int

    title: str

    username: str | None

    type: str

    updated_at: str

    dialog_started_at: str

    dialog_closed_at: str | None = None

    dialog_purge_at: str | None = None

    section: str | None = None

    section_title: str | None = None

    bin: str | None = None

    is_favorite: bool = False

    operator_mode: bool = False

    unread_count: int = 0

    last_message_text: str | None = None

    last_message_direction: str | None = None

    last_message_author: str | None = None

    last_message_has_attachments: bool = False

    last_message_attachment_kind: str | None = None

    employee_assessment_id: int | None = None

    employee_assessment_pending: bool = False

    employee_assessment_created_at: str | None = None


class DialogStatusResponse(BaseModel):

    status: str = Field(default="ok")

    chat_id: int

    dialog_id: int

    dialog_closed_at: str | None = None

    dialog_purge_at: str | None = None

    ai_enabled: bool = True

    employee_assessment_id: int | None = None

    employee_assessment_pending: bool = False





class NotificationResponse(BaseModel):

    type: str = Field(default="message")

    chat_id: int | None = None

    chat_title: str | None = None

    text: str

    created_at: str

    section: str | None = None

    section_title: str | None = None

    bin: str | None = None

    dialog_id: int | None = None





class DashboardSectionStat(BaseModel):

    section: Optional[str] = None

    title: str

    dialogs: int

    percentage: float





class DashboardTopQuestion(BaseModel):

    question: str

    count: int





class DashboardSectionQuestions(BaseModel):

    section: Optional[str] = None

    title: str

    questions: List[DashboardTopQuestion] = Field(default_factory=list)





class DashboardAgentStat(BaseModel):

    name: str

    dialogs: int

    messages: int

    avg_messages_per_dialog: float

    avg_response_time_minutes: float | None = None

    last_activity: Optional[str] = None

    avg_csat: float | None = None





class DashboardActivityPoint(BaseModel):

    date: str

    dialogs: int

    incoming_messages: int





class DashboardResponseTimeDialog(BaseModel):

    chat_id: int | None = None

    dialog_id: int | None = None

    author: str

    response_time_minutes: float





class DashboardTopBin(BaseModel):

    bin: str

    requests: int





class DashboardHeatmapPoint(BaseModel):

    day_of_week: int

    hour: int

    count: int



class DashboardDialogMetric(BaseModel):

    dialog_id: int

    bin: str | None

    is_open: bool

    is_ai_closed: bool

    response_time_minutes: float | None

    csat_rating: int | None = None

    ai_csat_rating: int | None = None
    rated_by: str | None = None
    operator_name: str | None = None



class CsatDistributionEntry(BaseModel):

    rating: int

    count: int





class DashboardSummaryResponse(BaseModel):

    total_dialogs: int

    open_dialogs: int

    closed_dialogs: int

    total_chats: int

    total_messages: int

    total_incoming_messages: int

    total_outgoing_messages: int

    ai_closed_dialogs: int = 0

    transferred_to_operator_dialogs: int = 0

    avg_messages_before_transfer: float | None = None

    ai_messages_count: int = 0

    requests_with_contract: int = 0

    requests_without_contract: int = 0

    recurring_requests_count: int = 0

    recurring_requests_percentage: float | None = None

    sla_violations_count: int = 0

    sla_compliance_percentage: float | None = None

    average_first_message_length: float | None = None

    average_messages_per_dialog: float

    avg_dialog_duration_minutes: float | None = None

    avg_response_time_minutes: float | None = None

    avg_response_time_seconds: float | None = None

    response_time_dialogs: List[DashboardResponseTimeDialog] = Field(default_factory=list)

    section_breakdown: List[DashboardSectionStat] = Field(default_factory=list)

    top_questions: List[DashboardTopQuestion] = Field(default_factory=list)

    questions_by_section: List[DashboardSectionQuestions] = Field(default_factory=list)

    agent_breakdown: List[DashboardAgentStat] = Field(default_factory=list)

    recent_activity: List[DashboardActivityPoint] = Field(default_factory=list)

    top_bins_without_contract: List[DashboardTopBin] = Field(default_factory=list)

    top_bins_with_contract: List[DashboardTopBin] = Field(default_factory=list)

    peak_load_heatmap: List[DashboardHeatmapPoint] = Field(default_factory=list)

    dialog_metrics: List[DashboardDialogMetric] = Field(default_factory=list)

    csat_average: float | None = None

    csat_count: int = 0

    csat_distribution: List[CsatDistributionEntry] = Field(default_factory=list)

    ai_csat_average: float | None = None

    ai_csat_count: int = 0

    ai_csat_distribution: List[CsatDistributionEntry] = Field(default_factory=list)

    updated_at: str





def require_api_token(

    x_api_token: str | None = Header(default=None, alias="X-Api-Token"),

    api_token: str | None = Query(default=None, alias="api_token"),

) -> None:

    if not API_TOKEN:

        raise HTTPException(status_code=503, detail="API token is not configured")

    token_to_check = x_api_token or api_token

    if token_to_check != API_TOKEN:

        raise HTTPException(status_code=403, detail="Invalid API token")
def require_onec_token(
    x_integration_token: str | None = Header(default=None, alias="X-Integration-Token")
) -> None:
    if not ONEC_INTEGRATION_TOKEN:
        raise HTTPException(status_code=503, detail="1C integration token is not configured")
    if x_integration_token != ONEC_INTEGRATION_TOKEN:
        raise HTTPException(status_code=403, detail="Invalid integration token")





def get_current_user(

    _: None = Depends(require_api_token),

    x_session_token: str | None = Header(default=None, alias="X-Session-Token"),

    session_token: str | None = Query(default=None, alias="session_token"),

) -> Dict[str, object]:

    token_to_check = x_session_token or session_token

    if not token_to_check:

        raise HTTPException(status_code=401, detail="Session token required")

    user = database.get_user_by_session(token_to_check)

    if not user:

        raise HTTPException(status_code=401, detail="Invalid session token")

    return _sanitize_user(user)





@router.get("/stream")

async def sse_endpoint(request: Request, current_user: Dict[str, object] = Depends(get_current_user)):

    user_id = current_user["id"]

    queue = asyncio.Queue()

    event_bus.connections[user_id].append(queue)

    logger.info("SSE client connected: user_id=%s", user_id)



    async def event_generator():

        try:

            while True:

                # If client closes connection, request.is_disconnected() will be true 

                # but run in next tick, so asyncio.wait allows us to detect it

                if await request.is_disconnected():

                    break

                try:

                    message = await asyncio.wait_for(queue.get(), timeout=1.0)

                    yield f"data: {json.dumps(message, ensure_ascii=False)}\n\n"

                except asyncio.TimeoutError:

                    # Keep-alive heartbeat

                    yield ": keepalive\n\n"

        finally:

            logger.info("SSE client disconnected: user_id=%s", user_id)

            if queue in event_bus.connections[user_id]:

                event_bus.connections[user_id].remove(queue)

            if not event_bus.connections[user_id]:

                del event_bus.connections[user_id]



    return StreamingResponse(
        event_generator(), 
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",  # Prevents Nginx/proxies from buffering SSE
        }
    )





def require_admin_or_moderator(

    current_user: Dict[str, object] = Depends(get_current_user),

) -> Dict[str, object]:

    if not database.is_admin_like(current_user["role"]):

        raise HTTPException(status_code=403, detail="Administrator role required")

    return current_user



def require_hr_access(

    current_user: Dict[str, object] = Depends(get_current_user),

) -> Dict[str, object]:

    if not database.can_manage_hr(str(current_user["role"])):

        raise HTTPException(status_code=403, detail="HR role required")

    return current_user





def _ensure_moderator_can_manage(current_user: Dict[str, object], target_user: Dict[str, object]) -> None:

    if current_user["role"] == database.ROLE_MODERATOR:

        if target_user["role"] != database.ROLE_OPERATOR:

            raise HTTPException(status_code=403, detail="\u041d\u0435\u0434\u043e\u0441\u0442\u0430\u0442\u043e\u0447\u043d\u043e \u043f\u0440\u0430\u0432 \u0434\u043b\u044f \u0443\u043f\u0440\u0430\u0432\u043b\u0435\u043d\u0438\u044f \u044d\u0442\u0438\u043c \u043f\u043e\u043b\u044c\u0437\u043e\u0432\u0430\u0442\u0435\u043b\u0435\u043c")





def _sanitize_user(user: Dict[str, object]) -> Dict[str, object]:

    user_id = user["id"]

    sections = user.get("sections") or database.get_user_sections(user_id)

    favorites = database.list_favorite_dialog_ids(user_id)

    raw_bins = user.get("bins")

    if raw_bins and all(isinstance(assignment, dict) for assignment in raw_bins):

        bins = raw_bins

    else:

        bins = database.get_user_bin_assignments(user_id)

    return {

        "id": user_id,

        "email": user["email"],

        "login": user.get("login", ""),

        "name": user["name"],

        "created_at": user["created_at"],

        "job_title": user.get("job_title", ""),

        "organization": user.get("organization", database.DEFAULT_EMPLOYEE_ORGANIZATION),

        "phone": user.get("phone", ""),

        "bio": user.get("bio", ""),

        "role": user.get("role", database.ROLE_OPERATOR),

        "is_approved": bool(user.get("is_approved", True)),

        "sections": sections,

        "bins": bins,

        "favorite_dialog_ids": favorites,

    }





@router.post("/auth/register", response_model=RegisterResponse)

def register_user(request: RegisterRequest, _: None = Depends(require_api_token)):

    logger.info("Registration attempt for email=%s", request.email)

    existing = database.find_user_by_email(request.email)

    if existing:

        logger.warning("Registration rejected: email=%s already exists", request.email)

        raise HTTPException(status_code=409, detail="User already exists")

    password_hash = hashlib.sha256(request.password.encode("utf-8")).hexdigest()

    try:

        created_user = database.create_user(

            request.email,

            request.name,

            password_hash,

            login=request.email,

            role=database.ROLE_OPERATOR,

            is_approved=False,

        )

    except ValueError as exc:

        logger.warning("Registration failed for email=%s: %s", request.email, exc)

        raise HTTPException(status_code=409, detail=str(exc)) from exc

    if created_user is None or created_user.get("id") is None:

        logger.error("Registration failed for email=%s: user not created", request.email)

        raise HTTPException(status_code=500, detail="Failed to create user")

    logger.info("Registration successful: email=%s, user_id=%s", request.email, created_user["id"])

    return RegisterResponse()





@router.post("/auth/login", response_model=AuthResponse)

def login_user(request: LoginRequest, _: None = Depends(require_api_token)):

    logger.info("Login attempt for identifier=%s", request.identifier)

    user = database.find_user_by_identifier(request.identifier)

    password_hash = hashlib.sha256(request.password.encode("utf-8")).hexdigest()

    if not user or user["password_hash"] != password_hash:

        logger.warning("Login failed: invalid credentials for identifier=%s", request.identifier)

        raise HTTPException(status_code=401, detail="Invalid credentials")

    if not user.get("is_approved", True):

        logger.warning("Login blocked: user_id=%s not approved", user["id"])

        raise HTTPException(status_code=403, detail="\u0410\u043a\u043a\u0430\u0443\u043d\u0442 \u043e\u0436\u0438\u0434\u0430\u0435\u0442 \u043f\u043e\u0434\u0442\u0432\u0435\u0440\u0436\u0434\u0435\u043d\u0438\u044f \u043c\u043e\u0434\u0435\u0440\u0430\u0442\u043e\u0440\u0430")

    token = database.create_session(user["id"])

    logger.info("Login successful: user_id=%s, role=%s", user["id"], user.get("role"))

    sanitized = _sanitize_user(user)

    return AuthResponse(token=token, user=UserResponse(**sanitized))





@router.get("/profile")

def get_profile(current_user: Dict[str, object] = Depends(get_current_user)):

    return UserResponse(**current_user)





class ProfileUpdateRequest(BaseModel):

    name: str | None = Field(default=None, min_length=2, max_length=100)

    job_title: str | None = Field(default=None, max_length=120)

    organization: str | None = Field(default=None, max_length=160)

    phone: str | None = Field(default=None, max_length=50)

    bio: str | None = Field(default=None, max_length=500)

    email: EmailStr | None = None





@router.put("/profile", response_model=UserResponse)

def update_profile(

    request: ProfileUpdateRequest,

    current_user: Dict[str, object] = Depends(get_current_user),

):

    payload = {

        "name": request.name if request.name is not None else current_user["name"],

        "job_title": request.job_title if request.job_title is not None else current_user.get("job_title", ""),

        "organization": request.organization if request.organization is not None else current_user.get("organization", database.DEFAULT_EMPLOYEE_ORGANIZATION),

        "phone": request.phone if request.phone is not None else current_user.get("phone", ""),

        "bio": request.bio if request.bio is not None else current_user.get("bio", ""),

        "email": request.email if request.email is not None else current_user.get("email"),

    }

    try:

        updated = database.update_user_profile(

            current_user["id"],

            name=payload["name"],

            job_title=payload["job_title"],

            organization=payload["organization"],

            phone=payload["phone"],

            bio=payload["bio"],

            email=payload["email"],

        )

    except ValueError as exc:

        raise HTTPException(status_code=409, detail=str(exc)) from exc

    return UserResponse(**updated)





@router.put("/profile/password", response_model=AuthResponse)

def change_password(

    request: PasswordChangeRequest,

    current_user: Dict[str, object] = Depends(get_current_user),

):

    current_hash = hashlib.sha256(request.current_password.encode("utf-8")).hexdigest()

    try:

        is_valid = database.verify_user_password(current_user["id"], current_hash)

    except ValueError as exc:

        raise HTTPException(status_code=404, detail=str(exc)) from exc

    if not is_valid:

        raise HTTPException(status_code=400, detail="\u0422\u0435\u043a\u0443\u0449\u0438\u0439 \u043f\u0430\u0440\u043e\u043b\u044c \u0443\u043a\u0430\u0437\u0430\u043d \u043d\u0435\u0432\u0435\u0440\u043d\u043e")

    new_hash = hashlib.sha256(request.new_password.encode("utf-8")).hexdigest()

    try:

        database.update_user_password(current_user["id"], new_hash)

    except ValueError as exc:

        raise HTTPException(status_code=404, detail=str(exc)) from exc

    fresh = database.get_user_by_id(current_user["id"])

    if fresh is None:

        raise HTTPException(status_code=404, detail="User not found")

    sanitized = _sanitize_user(fresh)

    token = database.create_session(current_user["id"])

    return AuthResponse(token=token, user=UserResponse(**sanitized))





@router.get("/users", response_model=List[UserResponse])

def list_users_admin(

    query: str | None = None,

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    users = database.list_users(query=query)

    return [UserResponse(**_sanitize_user(user)) for user in users]





@router.get("/users/pending", response_model=List[PendingUserResponse])

def list_pending_users(

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    pending = database.list_pending_users()

    return [

        PendingUserResponse(

            id=user["id"],

            email=user["email"],

            name=user["name"],

            created_at=user["created_at"],

        )

        for user in pending

    ]





@router.post("/users/{user_id}/approve", response_model=UserResponse)

def approve_user_registration(

    user_id: int,

    current_admin: Dict[str, object] = Depends(require_admin_or_moderator),

):

    user = database.get_user_by_id(user_id)

    if user is None:

        raise HTTPException(status_code=404, detail="User not found")

    _ensure_moderator_can_manage(current_admin, user)

    if user.get("is_approved", True):

        raise HTTPException(status_code=400, detail="\u0410\u043a\u043a\u0430\u0443\u043d\u0442 \u0443\u0436\u0435 \u043f\u043e\u0434\u0442\u0432\u0435\u0440\u0436\u0434\u0451\u043d")

    updated = database.set_user_approved(user_id, True)

    return UserResponse(**_sanitize_user(updated))





@router.post("/users/{user_id}/reject")

def reject_user_registration(

    user_id: int,

    current_admin: Dict[str, object] = Depends(require_admin_or_moderator),

):

    user = database.get_user_by_id(user_id)

    if user is None:

        raise HTTPException(status_code=404, detail="User not found")

    _ensure_moderator_can_manage(current_admin, user)

    if user.get("is_approved", True):

        raise HTTPException(status_code=400, detail="\u0410\u043a\u043a\u0430\u0443\u043d\u0442 \u0443\u0436\u0435 \u043f\u043e\u0434\u0442\u0432\u0435\u0440\u0436\u0434\u0451\u043d")

    database.delete_user(user_id)

    return {"status": "ok"}





@router.put("/users/{user_id}/role", response_model=UserResponse)

def set_user_role(

    user_id: int,

    request: RoleUpdateRequest,

    current_admin: Dict[str, object] = Depends(require_admin_or_moderator),

):

    desired_role = request.role.strip()

    if desired_role not in database.ALL_ROLES:

        raise HTTPException(status_code=400, detail="Unknown role")

    target_user = database.get_user_by_id(user_id)

    if target_user is None:

        raise HTTPException(status_code=404, detail="User not found")

    _ensure_moderator_can_manage(current_admin, target_user)

    if current_admin["role"] == database.ROLE_MODERATOR and desired_role != database.ROLE_OPERATOR:

        raise HTTPException(status_code=403, detail="\u041d\u0435\u0434\u043e\u0441\u0442\u0430\u0442\u043e\u0447\u043d\u043e \u043f\u0440\u0430\u0432 \u0434\u043b\u044f \u043d\u0430\u0437\u043d\u0430\u0447\u0435\u043d\u0438\u044f \u044d\u0442\u043e\u0439 \u0440\u043e\u043b\u0438")

    if user_id == current_admin["id"] and current_admin["role"] == database.ROLE_ADMIN and desired_role != database.ROLE_ADMIN:

        raise HTTPException(status_code=400, detail="\u0410\u0434\u043c\u0438\u043d\u0438\u0441\u0442\u0440\u0430\u0442\u043e\u0440 \u043d\u0435 \u043c\u043e\u0436\u0435\u0442 \u0441\u043d\u044f\u0442\u044c \u0441\u043e\u0431\u0441\u0442\u0432\u0435\u043d\u043d\u044b\u0435 \u043f\u0440\u0430\u0432\u0430")

    try:

        updated = database.update_user_role(user_id, desired_role)

    except ValueError as exc:

        raise HTTPException(status_code=404, detail=str(exc)) from exc

    return UserResponse(**_sanitize_user(updated))





@router.put("/users/{user_id}/password", response_model=UserResponse)

def admin_set_user_password(

    user_id: int,

    request: PasswordResetRequest,

    current_admin: Dict[str, object] = Depends(require_admin_or_moderator),

):

    target_user = database.get_user_by_id(user_id)

    if target_user is None:

        raise HTTPException(status_code=404, detail="User not found")

    _ensure_moderator_can_manage(current_admin, target_user)

    new_hash = hashlib.sha256(request.new_password.encode("utf-8")).hexdigest()

    try:

        database.update_user_password(user_id, new_hash)

    except ValueError as exc:

        raise HTTPException(status_code=404, detail=str(exc)) from exc

    fresh = database.get_user_by_id(user_id)

    if fresh is None:

        raise HTTPException(status_code=404, detail="User not found")

    return UserResponse(**_sanitize_user(fresh))





@router.put("/users/{user_id}/sections", response_model=UserResponse)

def set_user_sections_endpoint(

    user_id: int,

    request: SectionsUpdateRequest,

    current_admin: Dict[str, object] = Depends(require_admin_or_moderator),

):

    # Ensure target user exists

    user = database.get_user_by_id(user_id)

    if user is None:

        raise HTTPException(status_code=404, detail="User not found")

    _ensure_moderator_can_manage(current_admin, user)

    if user_id == current_admin["id"]:

        pass

    valid_ids = {section["id"] for section in database.SECTIONS}

    invalid = [section_id for section_id in request.sections if section_id not in valid_ids]

    if invalid:

        raise HTTPException(status_code=400, detail=f"Unknown sections: {', '.join(invalid)}")

    updated_sections = database.set_user_sections(user_id, request.sections)

    sanitized = _sanitize_user({**user, "sections": updated_sections})

    return UserResponse(**sanitized)





@router.put("/users/{user_id}/bins", response_model=UserResponse)

def set_user_bins_endpoint(

    user_id: int,

    request: BinsUpdateRequest,

    current_admin: Dict[str, object] = Depends(require_admin_or_moderator),

):

    user = database.get_user_by_id(user_id)

    if user is None:

        raise HTTPException(status_code=404, detail="User not found")

    _ensure_moderator_can_manage(current_admin, user)

    updated_bins = database.set_user_bins(user_id, request.bins, assigned_by=current_admin["id"])

    sanitized = _sanitize_user({**user, "bins": updated_bins})

    return UserResponse(**sanitized)





@router.delete("/users/{user_id}")

def delete_user_endpoint(

    user_id: int,

    current_admin: Dict[str, object] = Depends(require_admin_or_moderator),

):

    if user_id == current_admin["id"]:

        raise HTTPException(status_code=400, detail="Нельзя удалить собственный аккаунт")

    target_user = database.get_user_by_id(user_id)

    if target_user is None:

        raise HTTPException(status_code=404, detail="User not found")

    _ensure_moderator_can_manage(current_admin, target_user)

    try:

        database.delete_user(user_id)

    except ValueError as exc:

        raise HTTPException(status_code=404, detail=str(exc)) from exc

    return {"status": "ok"}





@router.get("/roles")

def list_roles(_: Dict[str, object] = Depends(require_admin_or_moderator)):

    return [

        {"id": role, "title": ROLE_LABELS.get(role, role)}

        for role in database.ALL_ROLES

    ]





@router.get("/sections")

def list_sections(_: Dict[str, object] = Depends(get_current_user)):

    return database.SECTIONS





@router.get("/bins", response_model=List[str])

def list_bins_endpoint(

    query: str | None = None,

    _: Dict[str, object] = Depends(get_current_user),

):

    return database.list_bins(query)





@router.delete("/bins/{bin_value}")

def delete_bin_endpoint(

    bin_value: str,

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    """Delete a BIN from the database."""

    removed = database.remove_bin(bin_value)

    if not removed:

        raise HTTPException(status_code=404, detail="\u0411\u0418\u041d \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d")

    return {"status": "ok"}





@router.get("/bins/detailed", response_model=List[BinDetailedResponse])

def list_bins_detailed_endpoint(

    query: str | None = None,

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    """Return BINs enriched with contract-availability information."""

    return [
        BinDetailedResponse(**item)
        for item in database.list_bin_contract_snapshots(query)
    ]


def _persist_bin_contract_result(bin_value: str, contract_data: Dict[str, Any]) -> bool:
    has_contract = bool(contract_data.get("has_contract", False))
    database.upsert_bin_contract_snapshot(
        bin_value,
        has_contract=has_contract,
        customer_legal_address=contract_data.get("customer_legal_address"),
        customer_bank_name_ru=contract_data.get("customer_bank_name_ru"),
        customer_name_ru=contract_data.get("customer_name_ru"),
    )
    if has_contract:
        database.remove_organization_without_contract(bin_value)
    else:
        database.add_organization_without_contract(
            customer_bin=bin_value,
            customer_legal_address=contract_data.get("customer_legal_address"),
            customer_bank_name_ru=contract_data.get("customer_bank_name_ru"),
            customer_name_ru=contract_data.get("customer_name_ru"),
        )
    return has_contract


@router.get("/bins/{bin_value}/info", response_model=BinDetailedResponse)

def get_bin_info_endpoint(

    bin_value: str,

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    """Return BIN details together with the GraphQL contract check result."""

    contract_data = contract_checker.check_customer_contracts(bin_value)
    has_contract = _persist_bin_contract_result(bin_value, contract_data)

    return BinDetailedResponse(

        bin=bin_value,

        has_contract=has_contract,

        customer_legal_address=contract_data.get("customer_legal_address"),

        customer_bank_name_ru=contract_data.get("customer_bank_name_ru"),
        customer_name_ru=contract_data.get("customer_name_ru"),

    )





@router.get("/bins/unassigned", response_model=List[UnassignedBinResponse])

def list_unassigned_bins_endpoint(

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    return [UnassignedBinResponse(**item) for item in database.list_unassigned_bins()]





@router.post("/bins/sync")

def sync_bins_with_contracts_endpoint(

    force: bool = False,

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    """Synchronize all BINs with the latest contract information."""

    result = database.sync_bins_with_contracts(force=force)

    return {

        "status": "ok",

        **result,

    }





@router.get("/bins/pending", response_model=List[UnassignedBinResponse])

def list_pending_bins_endpoint(

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    """Legacy alias for clients expecting the previous endpoint path."""

    return list_unassigned_bins_endpoint()





@router.get("/organizations/without-contracts", response_model=List[OrganizationWithoutContractResponse])

def list_organizations_without_contracts_endpoint(

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    """Возвращает список организаций без действующих договоров."""

    return [OrganizationWithoutContractResponse(**item) for item in database.list_organizations_without_contracts()]





@router.get("/faq")

def list_faq(_: Dict[str, object] = Depends(get_current_user)):

    return database.list_faq()





# ------------------ Reply Templates (Шаблоны быстрых ответов) ------------------





class HrTemplateRequest(BaseModel):
    title: str = Field(min_length=1, max_length=160)
    type: Literal["vacation", "advance", "sickLeave", "businessTrip", "certificate", "serviceLetter"]
    body: str = Field(min_length=1, max_length=5000)
    variables: List[str] = Field(default_factory=list)
    description: str = Field(default="", max_length=1000)
    status: Literal["active", "archived"] = "active"


class HrTemplateResponse(BaseModel):
    id: int
    title: str
    type: str
    description: str = ""
    body: str
    variables: List[str] = Field(default_factory=list)
    status: str
    created_by: int | None = None
    created_at: str
    updated_at: str


class HrSignaturePayload(BaseModel):
    signature: str = Field(min_length=1)
    signed_payload: str = Field(min_length=1)
    signed_at: str = Field(min_length=1)
    certificate_subject: str | None = None
    certificate_serial: str | None = None
    certificate_pem: str | None = None


class HrRequestSubmit(BaseModel):
    template_id: int
    values: Dict[str, Any] = Field(default_factory=dict)
    summary: str = Field(default="", max_length=1000)
    period: str = Field(default="", max_length=240)
    employee_signature: HrSignaturePayload


class HrDecisionRequest(BaseModel):
    status: Literal["approved", "rejected", "needsInfo"]
    comment: str = Field(default="", max_length=1000)
    hr_signature: HrSignaturePayload | None = None


class HrEmployeeOrganizationRequest(BaseModel):
    organization: Literal["ТОО Азия-Сервис"]


class HrRequestEventResponse(BaseModel):
    id: int
    request_id: int
    action: str
    actor_id: int | None = None
    actor_name: str = ""
    comment: str = ""
    created_at: str


class HrRequestResponse(BaseModel):
    id: int
    template_id: int | None = None
    template_title: str = ""
    type: str
    employee_id: int | None = None
    employee_name: str
    department: str = ""
    status: str
    values: Dict[str, Any] = Field(default_factory=dict)
    rendered_text: str
    summary: str = ""
    period: str = ""
    submitted_at: str
    updated_at: str
    decided_at: str | None = None
    decided_by: int | None = None
    decided_by_name: str | None = None
    decision_comment: str = ""
    employee_signature: HrSignaturePayload | None = None
    hr_signature: HrSignaturePayload | None = None
    events: List[HrRequestEventResponse] = Field(default_factory=list)


class HrEmployeeResponse(BaseModel):
    id: int
    email: EmailStr
    login: str
    name: str
    created_at: str
    job_title: str = ""
    organization: str = "ТОО Азия-Сервис"
    phone: str = ""
    bio: str = ""
    role: str
    is_approved: bool = True
    sections: List[str] = Field(default_factory=list)
    bins: List[BinAssignmentResponse] = Field(default_factory=list)
    favorite_dialog_ids: List[int] = []
    schedule: str = "09:00-18:00"


def _ensure_can_view_hr_request(current_user: Dict[str, object], hr_request: dict | None) -> dict:
    if hr_request is None:
        raise HTTPException(status_code=404, detail="HR request not found")
    role = str(current_user.get("role", ""))
    if database.can_manage_hr(role) or int(hr_request.get("employee_id") or 0) == int(current_user["id"]):
        return hr_request
    raise HTTPException(status_code=403, detail="Нет доступа к заявлению")


def _normalize_hr_statement(value: str) -> str:
    statement = re.sub(r"\s+", " ", (value or "").strip())
    if not statement:
        return ""
    if statement[-1] not in ".!?":
        statement += "."
    return statement


def _build_hr_statement(hr_request: dict) -> str:
    values = hr_request.get("values") if isinstance(hr_request.get("values"), dict) else {}
    explicit_statement = str(values.get("statement") or "").strip()
    if explicit_statement:
        return _normalize_hr_statement(explicit_statement)

    rendered_text = str(hr_request.get("rendered_text") or "").strip()
    if re.match(r"^я,\s", rendered_text, flags=re.IGNORECASE):
        return _normalize_hr_statement(rendered_text)

    employee_name = str(hr_request.get("employee_name") or "")
    period = str(hr_request.get("period") or "")
    reason = str(hr_request.get("summary") or "").strip()
    prefix = f"Я, {employee_name},"
    request_type = str(hr_request.get("type") or "")
    if request_type == "advance":
        return _normalize_hr_statement(f"{prefix} запрашиваю аванс {reason or period}")
    if request_type == "vacation":
        return _normalize_hr_statement(f"{prefix} прошу предоставить отпуск на период {period} в связи с {reason}")
    if request_type == "businessTrip":
        return _normalize_hr_statement(f"{prefix} прошу оформить командировку на период {period} с целью {reason}")
    if request_type == "certificate":
        return _normalize_hr_statement(f"{prefix} прошу подготовить справку с места работы для {reason or 'предоставления по месту требования'}")
    if request_type == "sickLeave":
        return _normalize_hr_statement(f"{prefix} прошу оформить отсутствие по болезни на период {period} по причине {reason}")
    return _normalize_hr_statement(f"{prefix} прошу рассмотреть заявление на период {period} по причине {reason}")


def _hr_document_context(hr_request: dict) -> dict[str, str]:
    values = hr_request.get("values") if isinstance(hr_request.get("values"), dict) else {}
    organization = str(values.get("organization") or "организации").strip()
    document_date = str(values.get("document_date") or "").strip() or datetime.now(timezone.utc).strftime("%d.%m.%Y")
    return {
        "to": "Директору",
        "organization": organization,
        "from": str(hr_request.get("employee_name") or ""),
        "position": str(hr_request.get("department") or ""),
        "title": "Заявление",
        "body": _build_hr_statement(hr_request),
        "date": document_date,
    }


def _hr_document_html(hr_request: dict) -> str:
    context = {key: html.escape(value) for key, value in _hr_document_context(hr_request).items()}
    return f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>{context['title']}</title>
  <style>
    @page {{ size: A4; margin: 24mm 20mm; }}
    body {{ font-family: Arial, sans-serif; color: #111827; font-size: 14pt; line-height: 1.55; }}
    .addressee {{ width: 44%; margin-left: auto; margin-bottom: 46mm; font-size: 12pt; line-height: 1.45; }}
    h1 {{ margin: 0 0 16mm; text-align: center; font-size: 18pt; letter-spacing: 0; }}
    .body {{ margin: 0 auto; max-width: 160mm; text-align: justify; }}
    .footer {{ display: table; width: 100%; margin-top: 38mm; font-size: 12pt; }}
    .footer > div {{ display: table-cell; width: 50%; vertical-align: top; }}
    .signature {{ text-align: right; }}
  </style>
</head>
<body>
  <div class="addressee">
    <div>{context['to']}</div>
    <div>{context['organization']}</div>
  </div>
  <h1>{context['title']}</h1>
  <div class="body">
    <p>{context['body']}</p>
  </div>
  <div class="footer">
    <div>{context['date']}</div>
    <div class="signature">________________ / {context['from']} /</div>
  </div>
</body>
</html>"""


def _find_reportlab_font() -> tuple[str, str | None]:
    env_font = os.getenv("HR_DOCUMENT_FONT", "").strip()
    candidates = [
        Path(env_font) if env_font else None,
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
    ]
    for candidate in candidates:
        if candidate is not None and candidate.is_file():
            return "HrDocumentFont", str(candidate)
    return "Helvetica", None


def _wrap_pdf_text(text: str, max_width: float, font_name: str, font_size: int) -> list[str]:
    from reportlab.pdfbase import pdfmetrics

    lines: list[str] = []
    for paragraph in text.splitlines() or [""]:
        words = paragraph.split()
        current = ""
        for word in words:
            probe = f"{current} {word}".strip()
            if pdfmetrics.stringWidth(probe, font_name, font_size) <= max_width:
                current = probe
                continue
            if current:
                lines.append(current)
            current = word
        lines.append(current)
    return lines


def _hr_document_pdf(hr_request: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas

    font_name, font_path = _find_reportlab_font()
    if font_path:
        pdfmetrics.registerFont(TTFont(font_name, font_path))

    context = _hr_document_context(hr_request)
    buffer = io.BytesIO()
    page = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin_x = 56
    y = height - 70

    page.setFont(font_name, 11)
    x_right = width - 245
    for line in [context["to"], context["organization"]]:
        page.drawString(x_right, y, line)
        y -= 18

    y -= 120
    page.setFont(font_name, 16)
    page.drawCentredString(width / 2, y, context["title"])

    y -= 46
    page.setFont(font_name, 12)
    for line in _wrap_pdf_text(context["body"], width - margin_x * 2, font_name, 12):
        if y < 110:
            page.showPage()
            page.setFont(font_name, 12)
            y = height - 70
        page.drawString(margin_x, y, line)
        y -= 19

    page.drawString(margin_x, 74, context["date"])
    page.drawRightString(width - margin_x, 74, f"________________ / {context['from']} /")
    page.save()
    return buffer.getvalue()


@router.get("/hr/employees", response_model=List[HrEmployeeResponse])
def list_hr_employees(
    query: str | None = None,
    _: Dict[str, object] = Depends(require_hr_access),
):
    employees = database.list_hr_employees(query=query)
    return [HrEmployeeResponse(**employee) for employee in employees]


@router.put("/hr/employees/{employee_id}/organization", response_model=HrEmployeeResponse)
def set_hr_employee_organization(
    employee_id: int,
    request: HrEmployeeOrganizationRequest,
    _: Dict[str, object] = Depends(require_hr_access),
):
    if database.get_user_by_id(employee_id) is None:
        raise HTTPException(status_code=404, detail="User not found")
    try:
        employee = database.update_user_organization(employee_id, request.organization)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return HrEmployeeResponse(**{**employee, "schedule": "09:00-18:00"})


@router.get("/hr/templates", response_model=List[HrTemplateResponse])
def list_hr_templates(current_user: Dict[str, object] = Depends(get_current_user)):
    templates = database.list_hr_templates(
        active_only=not database.can_manage_hr(str(current_user.get("role", "")))
    )
    return [HrTemplateResponse(**template) for template in templates]


@router.post("/hr/templates", response_model=HrTemplateResponse)
def create_hr_template(
    request: HrTemplateRequest,
    current_user: Dict[str, object] = Depends(require_hr_access),
):
    try:
        template = database.create_hr_template(
            title=request.title,
            type=request.type,
            body=request.body,
            variables=request.variables,
            description=request.description,
            status=request.status,
            created_by=int(current_user["id"]),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return HrTemplateResponse(**template)


@router.put("/hr/templates/{template_id}", response_model=HrTemplateResponse)
def update_hr_template(
    template_id: int,
    request: HrTemplateRequest,
    _: Dict[str, object] = Depends(require_hr_access),
):
    try:
        template = database.update_hr_template(
            template_id,
            title=request.title,
            type=request.type,
            body=request.body,
            variables=request.variables,
            description=request.description,
            status=request.status,
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return HrTemplateResponse(**template)


@router.get("/hr/requests", response_model=List[HrRequestResponse])
def list_hr_requests(current_user: Dict[str, object] = Depends(get_current_user)):
    employee_id = None if database.can_manage_hr(str(current_user.get("role", ""))) else int(current_user["id"])
    requests = database.list_hr_requests(employee_id=employee_id)
    return [HrRequestResponse(**request) for request in requests]


@router.get("/hr/requests/{request_id}/document.doc")
def download_hr_request_doc(
    request_id: int,
    current_user: Dict[str, object] = Depends(require_hr_access),
):
    hr_request = _ensure_can_view_hr_request(current_user, database.get_hr_request(int(request_id)))
    content = _hr_document_html(hr_request).encode("utf-8")
    headers = {"Content-Disposition": f'attachment; filename="hr-request-{request_id}.doc"'}
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/msword; charset=utf-8",
        headers=headers,
    )


@router.get("/hr/requests/{request_id}/document.pdf")
def download_hr_request_pdf(
    request_id: int,
    current_user: Dict[str, object] = Depends(require_hr_access),
):
    hr_request = _ensure_can_view_hr_request(current_user, database.get_hr_request(int(request_id)))
    content = _hr_document_pdf(hr_request)
    headers = {"Content-Disposition": f'attachment; filename="hr-request-{request_id}.pdf"'}
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/pdf",
        headers=headers,
    )


@router.post("/hr/requests", response_model=HrRequestResponse)
def create_hr_request(
    request: HrRequestSubmit,
    current_user: Dict[str, object] = Depends(get_current_user),
):
    try:
        values = dict(request.values or {})
        values["organization"] = str(current_user.get("organization") or database.DEFAULT_EMPLOYEE_ORGANIZATION)
        hr_request = database.create_hr_request(
            template_id=request.template_id,
            employee_id=int(current_user["id"]),
            employee_name=str(current_user.get("name") or current_user.get("login") or ""),
            department=str(current_user.get("job_title") or ""),
            values=values,
            summary=request.summary,
            period=request.period,
            employee_signature=request.employee_signature.model_dump(),
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return HrRequestResponse(**hr_request)


@router.post("/hr/requests/{request_id}/decision", response_model=HrRequestResponse)
def decide_hr_request(
    request_id: int,
    request: HrDecisionRequest,
    current_user: Dict[str, object] = Depends(require_hr_access),
):
    try:
        if request.status in {"approved", "rejected"} and request.hr_signature is None:
            raise HTTPException(status_code=400, detail="HR signature is required for approval or rejection")
        hr_request = database.decide_hr_request(
            request_id,
            status=request.status,
            decided_by=int(current_user["id"]),
            decided_by_name=str(current_user.get("name") or current_user.get("login") or ""),
            comment=request.comment,
            hr_signature=request.hr_signature.model_dump() if request.hr_signature else None,
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return HrRequestResponse(**hr_request)


class ReplyTemplateRequest(BaseModel):

    title: str = Field(min_length=1, max_length=100)

    text: str = Field(min_length=1, max_length=2000)

    section: str | None = None

    sort_order: int = 0





class ReplyTemplateResponse(BaseModel):

    id: int

    title: str

    text: str

    section: str | None = None

    section_title: str | None = None

    sort_order: int = 0

    created_by: int | None = None

    created_at: str





def _enrich_template(template: dict) -> dict:

    """Add section_title from SECTIONS list."""

    section_id = template.get("section")

    section_title = None

    if section_id:

        section = next((s for s in database.SECTIONS if s["id"] == section_id), None)

        if section:

            section_title = section["title"]

    return {**template, "section_title": section_title}





@router.get("/reply-templates", response_model=List[ReplyTemplateResponse])

def list_reply_templates(

    section: str | None = None,

    _: Dict[str, object] = Depends(get_current_user),

):

    templates = database.list_reply_templates(section=section)

    return [ReplyTemplateResponse(**_enrich_template(t)) for t in templates]





@router.post("/reply-templates", response_model=ReplyTemplateResponse)

def create_reply_template(

    request: ReplyTemplateRequest,

    current_user: Dict[str, object] = Depends(require_admin_or_moderator),

):

    template = database.create_reply_template(

        title=request.title,

        text=request.text,

        section=request.section,

        sort_order=request.sort_order,

        created_by=current_user["id"],

    )

    logger.info(

        "Reply template created: id=%s, title=%s, by user_id=%s",

        template["id"], template["title"], current_user["id"],

    )

    return ReplyTemplateResponse(**_enrich_template(template))





@router.put("/reply-templates/{template_id}", response_model=ReplyTemplateResponse)

def update_reply_template(

    template_id: int,

    request: ReplyTemplateRequest,

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    try:

        template = database.update_reply_template(

            template_id,

            title=request.title,

            text=request.text,

            section=request.section,

            sort_order=request.sort_order,

        )

    except ValueError as exc:

        raise HTTPException(status_code=404, detail=str(exc)) from exc

    logger.info("Reply template updated: id=%s", template_id)

    return ReplyTemplateResponse(**_enrich_template(template))





@router.delete("/reply-templates/{template_id}")

def delete_reply_template(

    template_id: int,

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    deleted = database.delete_reply_template(template_id)

    if not deleted:

        raise HTTPException(status_code=404, detail="Шаблон не найден")

    logger.info("Reply template deleted: id=%s", template_id)

    return {"status": "ok"}





@router.get("/analytics/dashboard", response_model=DashboardSummaryResponse)

def dashboard_summary(

    operator_id: int | None = Query(default=None),

    start_date: date | None = Query(default=None),

    end_date: date | None = Query(default=None),

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    summary = database.get_dashboard_summary(

        operator_id=operator_id,

        start_date=start_date,

        end_date=end_date,

    )

    return DashboardSummaryResponse(**summary)





@router.get("/analytics/export")

def export_dashboard(

    operator_id: int | None = Query(default=None),

    start_date: date | None = Query(default=None),

    end_date: date | None = Query(default=None),

    format: str = Query(default="xlsx"),

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    """Generate an Excel or PDF report from dashboard data."""

    summary = database.get_dashboard_summary(

        operator_id=operator_id,

        start_date=start_date,

        end_date=end_date,

    )



    def _fmt(value, suffix=""):

        if value is None:

            return "—"

        if isinstance(value, float):

            return f"{value:.1f}{suffix}"

        return f"{value}{suffix}"



    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")



    if format == "pdf":

        return _build_pdf_report(summary, _fmt, now_str)



    return _build_xlsx_report(summary, _fmt, now_str)





def _build_xlsx_report(summary: dict, _fmt, now_str: str):

    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    from openpyxl.worksheet.table import Table, TableStyleInfo

    from openpyxl.utils import get_column_letter



    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)

    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")

    header_align = Alignment(horizontal="center", vertical="center")

    thin_border = Border(

        left=Side(style="thin", color="E2E8F0"),

        right=Side(style="thin", color="E2E8F0"),

        top=Side(style="thin", color="E2E8F0"),

        bottom=Side(style="thin", color="E2E8F0"),

    )



    def _style_header(ws, headers: list[str]):

        for col_idx, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col_idx, value=header)

            cell.font = header_font

            cell.fill = header_fill

            cell.alignment = header_align

            cell.border = thin_border



    def _auto_width(ws):

        for column_cells in ws.columns:

            max_len = 0

            col_letter = column_cells[0].column_letter

            for cell in column_cells:

                try:

                    if cell.value:

                        max_len = max(max_len, len(str(cell.value)))

                except Exception:

                    pass

            ws.column_dimensions[col_letter].width = min(max_len + 5, 60)



    def _add_table(ws, display_name: str, num_rows: int, num_cols: int):

        if num_rows <= 1:

            return

        ref = f"A1:{get_column_letter(num_cols)}{num_rows}"

        tab = Table(displayName=display_name.replace(" ", "_").replace("(", "").replace(")", ""), ref=ref)

        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,

                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)

        tab.tableStyleInfo = style

        ws.add_table(tab)



    # ── Sheet 1: Обзор ──

    ws = wb.active

    ws.title = "Обзор"

    headers1 = ["Метрика", "Значение"]

    _style_header(ws, headers1)

    overview_rows = [

        ("Всего диалогов", summary.get("total_dialogs", 0)),

        ("Открытых", summary.get("open_dialogs", 0)),

        ("Закрытых", summary.get("closed_dialogs", 0)),

        ("Всего чатов", summary.get("total_chats", 0)),

        ("Всего сообщений", summary.get("total_messages", 0)),

        ("Входящих", summary.get("total_incoming_messages", 0)),

        ("\u0418\u0441\u0445\u043e\u0434\u044f\u0449\u0438\u0445", summary.get("total_outgoing_messages", 0)),

        ("Ср. сообщений/диалог", _fmt(summary.get("average_messages_per_dialog"))),

        ("Ср. время ответа (мин)", _fmt(summary.get("avg_response_time_minutes"), " мин")),

        ("Ср. длительность диалога (мин)", _fmt(summary.get("avg_dialog_duration_minutes"), " мин")),

        ("Решено ботом (AI)", summary.get("ai_closed_dialogs", 0)),

        ("Переведено оператору", summary.get("transferred_to_operator_dialogs", 0)),

        ("Сообщений от AI", summary.get("ai_messages_count", 0)),

        ("Нарушений SLA", summary.get("sla_violations_count", 0)),

        ("SLA (% соблюдения)", _fmt(summary.get("sla_compliance_percentage"), "%")),

        ("Повторных обращений", summary.get("recurring_requests_count", 0)),

        ("% повторных", _fmt(summary.get("recurring_requests_percentage"), "%")),

        ("С договором", summary.get("requests_with_contract", 0)),

        ("Без договора", summary.get("requests_without_contract", 0)),

    ]

    for row_idx, (metric, value) in enumerate(overview_rows, 2):

        ws.cell(row=row_idx, column=1, value=metric).border = thin_border

        ws.cell(row=row_idx, column=2, value=value).border = thin_border

    _add_table(ws, "OverviewTable", len(overview_rows)+1, 2)

    _auto_width(ws)



    # ── Sheet 2: Операторы ──

    ws2 = wb.create_sheet("Операторы")

    headers2 = ["Сотрудник", "Обращений", "Сообщений", "Ср. сообщ./обр.", "Ср. время ответа (мин)", "Последняя активность"]

    _style_header(ws2, headers2)

    agents = summary.get("agent_breakdown", [])

    for row_idx, agent in enumerate(agents, 2):

        ws2.cell(row=row_idx, column=1, value=agent.get("name", "")).border = thin_border

        ws2.cell(row=row_idx, column=2, value=agent.get("dialogs", 0)).border = thin_border

        ws2.cell(row=row_idx, column=3, value=agent.get("messages", 0)).border = thin_border

        ws2.cell(row=row_idx, column=4, value=_fmt(agent.get("avg_messages_per_dialog"))).border = thin_border

        ws2.cell(row=row_idx, column=5, value=_fmt(agent.get("avg_response_time_minutes"))).border = thin_border

        ws2.cell(row=row_idx, column=6, value=agent.get("last_activity", "—")).border = thin_border

    _add_table(ws2, "AgentsTable", len(agents)+1, 6)

    _auto_width(ws2)



    # ── Sheet 3: Разделы ──

    ws3 = wb.create_sheet("\u0420\u0430\u0437\u0434\u0435\u043b\u044b")

    headers3 = ["Раздел", "Диалогов", "Доля (%)"]

    _style_header(ws3, headers3)

    sections = summary.get("section_breakdown", [])

    for row_idx, section in enumerate(sections, 2):

        ws3.cell(row=row_idx, column=1, value=section.get("title", "")).border = thin_border

        ws3.cell(row=row_idx, column=2, value=section.get("dialogs", 0)).border = thin_border

        ws3.cell(row=row_idx, column=3, value=_fmt(section.get("percentage"), "%")).border = thin_border

    _add_table(ws3, "SectionsTable", len(sections)+1, 3)

    _auto_width(ws3)



    # ── Sheet 4: Активность ──

    ws4 = wb.create_sheet("Активность")

    headers4 = ["Дата", "Диалогов", "Входящих сообщений"]

    _style_header(ws4, headers4)

    activity = summary.get("recent_activity", [])

    for row_idx, point in enumerate(activity, 2):

        ws4.cell(row=row_idx, column=1, value=point.get("date", "")).border = thin_border

        ws4.cell(row=row_idx, column=2, value=point.get("dialogs", 0)).border = thin_border

        ws4.cell(row=row_idx, column=3, value=point.get("incoming_messages", 0)).border = thin_border

    _add_table(ws4, "ActivityTable", len(activity)+1, 3)

    _auto_width(ws4)



    # -- Sheet 5: BINs with contracts --

    ws5 = wb.create_sheet("\u0411\u0418\u041d\u044b (\u0441 \u0434\u043e\u0433\u043e\u0432\u043e\u0440\u0430\u043c\u0438)")

    headers5 = ["\u0411\u0418\u041d", "\u041e\u0431\u0440\u0430\u0449\u0435\u043d\u0438\u0439"]

    _style_header(ws5, headers5)

    bins_with = summary.get("top_bins_with_contract", [])

    for row_idx, bin_row in enumerate(bins_with, 2):

        ws5.cell(row=row_idx, column=1, value=bin_row.get("bin", "")).border = thin_border

        ws5.cell(row=row_idx, column=2, value=bin_row.get("requests", 0)).border = thin_border

    _add_table(ws5, "BinsWithContract", len(bins_with)+1, 2)

    _auto_width(ws5)



    # -- Sheet 6: BINs without contracts --

    ws6 = wb.create_sheet("\u0411\u0418\u041d\u044b (\u0431\u0435\u0437 \u0434\u043e\u0433\u043e\u0432\u043e\u0440\u043e\u0432)")

    headers6 = ["\u0411\u0418\u041d", "\u041e\u0431\u0440\u0430\u0449\u0435\u043d\u0438\u0439"]

    _style_header(ws6, headers6)

    bins_without = summary.get("top_bins_without_contract", [])

    for row_idx, bin_row in enumerate(bins_without, 2):

        ws6.cell(row=row_idx, column=1, value=bin_row.get("bin", "")).border = thin_border

        ws6.cell(row=row_idx, column=2, value=bin_row.get("requests", 0)).border = thin_border

    _add_table(ws6, "BinsWithoutContract", len(bins_without)+1, 2)

    _auto_width(ws6)



    # ── Sheet 7: Нагрузка по часам ──

    ws7 = wb.create_sheet("Нагрузка по часам")

    headers7 = ["День", "Час", "Кол-во обращений"]

    day_names = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]

    _style_header(ws7, headers7)

    heatmap = [e for e in summary.get("peak_load_heatmap", []) if e.get("count", 0) > 0]

    for row_idx, entry in enumerate(heatmap, 2):

        day_idx = entry.get("day", 0)

        ws7.cell(row=row_idx, column=1, value=day_names[day_idx] if 0 <= day_idx < 7 else str(day_idx)).border = thin_border

        ws7.cell(row=row_idx, column=2, value=f"{entry.get('hour', 0)}:00").border = thin_border

        ws7.cell(row=row_idx, column=3, value=entry.get("count", 0)).border = thin_border

    _add_table(ws7, "HeatmapTable", len(heatmap)+1, 3)

    _auto_width(ws7)



    # ── Save to buffer and return ──

    buf = io.BytesIO()

    wb.save(buf)

    buf.seek(0)



    filename = f"report_{now_str}.xlsx"

    return StreamingResponse(

        buf,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={"Content-Disposition": f'attachment; filename="{filename}"'},

    )







def _build_pdf_report(summary: dict, _fmt, now_str: str):

    from reportlab.lib import colors

    from reportlab.lib.pagesizes import A4

    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    from reportlab.lib.units import mm

    from reportlab.platypus import (

        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image

    )

    from reportlab.pdfbase import pdfmetrics

    from reportlab.pdfbase.ttfonts import TTFont

    import matplotlib

    from pathlib import Path



    # Load DejaVuSans from matplotlib to support Cyrillic

    mpl_data_dir = Path(matplotlib.get_data_path())

    font_path = mpl_data_dir / "fonts" / "ttf" / "DejaVuSans.ttf"

    if font_path.exists():

        pdfmetrics.registerFont(TTFont("DejaVu", str(font_path)))

        font_name = "DejaVu"

    else:

        font_name = "Helvetica" # Fallback, though Cyrillic will fail



    buf = io.BytesIO()

    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15 * mm, bottomMargin=15 * mm)

    styles = getSampleStyleSheet()



    # ── Colours ──

    brand = colors.HexColor("#6366F1")

    header_text_color = colors.white

    row_alt = colors.HexColor("#F8FAFC")

    border_color = colors.HexColor("#E2E8F0")



    title_style = ParagraphStyle(

        "PDFTitle", parent=styles["Title"], fontName=font_name, fontSize=18,

        textColor=brand, spaceAfter=4,

    )

    section_style = ParagraphStyle(

        "PDFSection", parent=styles["Heading2"], fontName=font_name, fontSize=13,

        textColor=brand, spaceBefore=14, spaceAfter=6,

    )

    normal = styles["Normal"]

    normal.fontName = font_name



    day_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]



    def _pdf_table(headers: list[str], rows: list[list], col_widths=None):

        data = [headers] + rows

        t = Table(data, colWidths=col_widths, repeatRows=1)

        style_cmds = [

            ("BACKGROUND", (0, 0), (-1, 0), brand),

            ("TEXTCOLOR", (0, 0), (-1, 0), header_text_color),

            ("FONTNAME", (0, 0), (-1, -1), font_name),

            ("FONTSIZE", (0, 0), (-1, 0), 9),

            ("FONTSIZE", (0, 1), (-1, -1), 8),

            ("ALIGN", (0, 0), (-1, -1), "LEFT"),

            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

            ("GRID", (0, 0), (-1, -1), 0.4, border_color),

            ("TOPPADDING", (0, 0), (-1, -1), 4),

            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),

            ("LEFTPADDING", (0, 0), (-1, -1), 6),

            ("RIGHTPADDING", (0, 0), (-1, -1), 6),

        ]

        for i in range(1, len(data)):

            if i % 2 == 0:

                style_cmds.append(("BACKGROUND", (0, i), (-1, i), row_alt))

        t.setStyle(TableStyle(style_cmds))

        return t



    def _create_pie_chart(sections):

        try:

            import matplotlib

            matplotlib.use('Agg')

            import matplotlib.pyplot as plt

            if not sections:
                return None

            fig, ax = plt.subplots(figsize=(5, 3.5))

            

            sorted_sec = sorted(sections, key=lambda x: x.get("dialogs", 0), reverse=True)

            top_sec = sorted_sec[:8]

            other_sum = sum(s.get("dialogs", 0) for s in sorted_sec[8:])

            labels = [s.get("title", "Unknown")[:20] for s in top_sec]

            sizes = [s.get("dialogs", 0) for s in top_sec]

            if other_sum > 0:

                labels.append("\u041e\u0441\u0442\u0430\u043b\u044c\u043d\u044b\u0435")

                sizes.append(other_sum)

            

            if sum(sizes) == 0:
                return None

            

            ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90, colors=plt.cm.Pastel1.colors)

            ax.axis('equal')

            plt.title("\u0420\u0430\u0437\u0434\u0435\u043b\u044b (\u0434\u043e\u043b\u044f \u043e\u0431\u0440\u0430\u0449\u0435\u043d\u0438\u0439)")

            

            img_buf = io.BytesIO()

            plt.savefig(img_buf, format='png', bbox_inches='tight', dpi=150)

            plt.close(fig)

            img_buf.seek(0)

            return img_buf

        except Exception as e:

            print("Matplotlib error:", e)

            return None



    elements = []



    # ── Title ──

    elements.append(Paragraph(f"Аналитический отчёт — {now_str}", title_style))

    elements.append(Spacer(1, 6))



    # ── 1. Обзор ──

    elements.append(Paragraph("Обзор", section_style))

    overview_rows = [

        ["Всего диалогов", str(summary.get("total_dialogs", 0))],

        ["Открытых", str(summary.get("open_dialogs", 0))],

        ["Закрытых", str(summary.get("closed_dialogs", 0))],

        ["Всего сообщений", str(summary.get("total_messages", 0))],

        ["Входящих", str(summary.get("total_incoming_messages", 0))],

        ["\u0418\u0441\u0445\u043e\u0434\u044f\u0449\u0438\u0445", str(summary.get("total_outgoing_messages", 0))],

        ["Ср. время ответа", _fmt(summary.get("avg_response_time_minutes"), " мин")],

        ["Решено ботом (AI)", str(summary.get("ai_closed_dialogs", 0))],

        ["Переведено оператору", str(summary.get("transferred_to_operator_dialogs", 0))],

        ["SLA (%)", _fmt(summary.get("sla_compliance_percentage"), "%")],

        ["Нарушений SLA", str(summary.get("sla_violations_count", 0))],

        ["С договором", str(summary.get("requests_with_contract", 0))],

        ["Без договора", str(summary.get("requests_without_contract", 0))],

    ]

    elements.append(_pdf_table(["Метрика", "Значение"], overview_rows, col_widths=[120 * mm, 50 * mm]))



    # ── 2. Операторы ──

    agents = summary.get("agent_breakdown", [])

    if agents:

        elements.append(Paragraph("Операторы", section_style))

        agent_rows = [

            [

                a.get("name", ""),

                str(a.get("dialogs", 0)),

                str(a.get("messages", 0)),

                _fmt(a.get("avg_response_time_minutes")),

            ]

            for a in agents

        ]

        elements.append(_pdf_table(

            ["Сотрудник", "Обращ.", "Сообщ.", "Ср. ответ"],

            agent_rows,

        ))



    # ── 3. Разделы ──

    sections = summary.get("section_breakdown", [])

    if sections:

        elements.append(Paragraph("\u0420\u0430\u0437\u0434\u0435\u043b\u044b", section_style))

        chart_buf = _create_pie_chart(sections)

        if chart_buf:

            elements.append(Image(chart_buf, width=100*mm, height=70*mm))

            elements.append(Spacer(1, 6))

            

        section_rows = [

            [s.get("title", ""), str(s.get("dialogs", 0)), _fmt(s.get("percentage"), "%")]

            for s in sections

        ]

        elements.append(_pdf_table(["Раздел", "Диалогов", "Доля"], section_rows))



    # ── 4. Активность ──

    activity = summary.get("recent_activity", [])

    if activity:

        elements.append(Paragraph("Активность по дням", section_style))

        act_rows = [

            [p.get("date", ""), str(p.get("dialogs", 0)), str(p.get("incoming_messages", 0))]

            for p in activity

        ]

        elements.append(_pdf_table(["Дата", "Диалогов", "Входящих"], act_rows))



    # -- 5. BINs with contracts --

    bins_with = summary.get("top_bins_with_contract", [])

    if bins_with:

        elements.append(Paragraph("\u0411\u0418\u041d\u044b (\u0441 \u0434\u043e\u0433\u043e\u0432\u043e\u0440\u0430\u043c\u0438)", section_style))

        bin_rows = [

            [b.get("bin", "—"), str(b.get("requests", 0))]

            for b in bins_with

        ]

        elements.append(_pdf_table(["\u0411\u0418\u041d", "\u041e\u0431\u0440\u0430\u0449\u0435\u043d\u0438\u0439"], bin_rows))



    # -- 6. BINs without contracts --

    bins_without = summary.get("top_bins_without_contract", [])

    if bins_without:

        elements.append(Paragraph("\u0411\u0418\u041d\u044b (\u0431\u0435\u0437 \u0434\u043e\u0433\u043e\u0432\u043e\u0440\u043e\u0432)", section_style))

        bin_rows2 = [

            [b.get("bin", "—"), str(b.get("requests", 0))]

            for b in bins_without

        ]

        elements.append(_pdf_table(["\u0411\u0418\u041d", "\u041e\u0431\u0440\u0430\u0449\u0435\u043d\u0438\u0439"], bin_rows2))



    # ── 7. Нагрузка по часам ──

    heatmap = summary.get("peak_load_heatmap", [])

    heat_mapped = [e for e in heatmap if e.get("count", 0) > 0]

    if heat_mapped:

        elements.append(Paragraph("Нагрузка по часам", section_style))

        heat_rows = [

            [

                day_names[e.get("day", 0)] if 0 <= e.get("day", 0) < 7 else str(e.get("day", 0)),

                f"{e.get('hour', 0)}:00",

                str(e.get("count", 0)),

            ]

            for e in heat_mapped

        ]

        elements.append(_pdf_table(["День", "Час", "Обращений"], heat_rows))



    doc.build(elements)

    buf.seek(0)



    filename = f"report_{now_str}.pdf"

    return StreamingResponse(

        buf,

        media_type="application/pdf",

        headers={"Content-Disposition": f'attachment; filename="{filename}"'},

    )



@router.get("/chats", response_model=List[ChatResponse])

def list_chats(

    favorite_only: bool = False,

    bin_query: str | None = None,

    current_user: Dict[str, object] = Depends(get_current_user),

):

    chats = database.list_chats_for_user(

        current_user["id"],

        current_user["role"],

        favorite_only=favorite_only,

        bin_query=bin_query,

    )

    enriched: List[ChatResponse] = []



    def _normalize(value: object, *, fallback: Optional[str] = None) -> str:

        if value is None:

            return fallback or datetime.now(timezone.utc).isoformat()

        if isinstance(value, datetime):

            return value.isoformat()

        return str(value)



    for chat in chats:

        section_id = chat.get("section")

        section_title = None

        if section_id:

            section = next((s for s in database.SECTIONS if s["id"] == section_id), None)

            if section:

                section_title = section["title"]

        enriched.append(

            ChatResponse(

                chat_id=int(chat["chat_id"]),

                dialog_id=int(chat["dialog_id"]),

                title=str(chat["title"]),

                username=chat.get("username"),

                type=str(chat["type"]),

                updated_at=_normalize(chat.get("updated_at")),

                dialog_started_at=_normalize(chat.get("dialog_started_at")),

                dialog_closed_at=_normalize(chat.get("dialog_closed_at")) if chat.get("dialog_closed_at") else None,

                dialog_purge_at=_normalize(chat.get("dialog_purge_at")) if chat.get("dialog_purge_at") else None,

                section=section_id,

                section_title=section_title,

                bin=chat.get("bin"),

                is_favorite=bool(chat.get("is_favorite")),

                operator_mode=bool(chat.get("operator_mode")),

                unread_count=int(chat.get("unread_count") or 0),

                last_message_text=chat.get("last_message_text"),

                last_message_direction=chat.get("last_message_direction"),

                last_message_author=chat.get("last_message_author"),

                last_message_has_attachments=bool(chat.get("last_message_has_attachments")),

                last_message_attachment_kind=chat.get("last_message_attachment_kind"),
                employee_assessment_id=chat.get("employee_assessment_id"),
                employee_assessment_pending=bool(chat.get("employee_assessment_pending")),
                employee_assessment_created_at=chat.get("employee_assessment_created_at"),
            )

        )

    return enriched





async def _store_uploaded_media(
    request: Request,
    upload: UploadFile | None,
    original_name: str | None,
    mime_type: str | None,
):
    try:
        if upload is not None:
            try:
                return media_service.ingest_upload(
                    upload.file,
                    original_name=upload.filename or original_name,
                    claimed_mime_type=upload.content_type or mime_type,
                )
            finally:
                await upload.close()

        # Use streaming to avoid loading entire file into memory
        fd, temp_name = tempfile.mkstemp(prefix="mtb-stream-", suffix=".bin")
        temp_path = Path(temp_name)
        try:
            with os.fdopen(fd, "wb") as tmp:
                async for chunk in request.stream():
                    tmp.write(chunk)
            
            resolved_name = original_name or request.headers.get("X-File-Name") or request.query_params.get("file_name")
            resolved_mime = (
                mime_type
                or request.headers.get("X-Mime-Type")
                or request.query_params.get("mime_type")
                or request.headers.get("Content-Type")
            )
            
            with open(temp_path, "rb") as f:
                return media_service.ingest_upload(
                    f,
                    original_name=resolved_name,
                    claimed_mime_type=resolved_mime,
                )
        finally:
            temp_path.unlink(missing_ok=True)
    except MediaValidationError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.message) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


def _absolute_api_url(request: Request, path: str) -> str:
    base_url = MEDIA_PUBLIC_BASE_URL or str(request.base_url).rstrip('/')
    if not MEDIA_PUBLIC_BASE_URL:
        parsed = urlsplit(base_url)
        host = (parsed.hostname or '').lower()
        if host in ('localhost', '::1', '[::1]'):
            netloc = '127.0.0.1'
            if parsed.port:
                netloc = f"{netloc}:{parsed.port}"
            base_url = urlunsplit((parsed.scheme or 'http', netloc, parsed.path, parsed.query, parsed.fragment)).rstrip('/')
    return f"{base_url}/{path.lstrip('/')}"




def _attachment_response_from_record(record: Dict[str, object], request: Request, for_onec: bool = False) -> AttachmentResponse:
    media_id = int(record["media_id"])

    url = media_service.build_signed_media_url(
        base_url=_absolute_api_url(request, f"/api/media/{media_id}"),
        media_id=media_id,
    )
    preview_url = None
    if str(record.get("kind") or "") == "image":
        preview_url = media_service.build_signed_media_url(
            base_url=_absolute_api_url(request, f"/api/media/{media_id}/preview"),
            media_id=media_id,
            variant="preview",
        )

    return AttachmentResponse(
        id=int(record["id"]),
        media_id=media_id,
        kind=str(record["kind"]),
        url=url,
        preview_url=preview_url,
        mime_type=str(record["mime_type"]),
        size_bytes=int(record["size_bytes"]),
        original_name=str(record["original_name"]),
        width=int(record["width"]) if record.get("width") is not None else None,
        height=int(record["height"]) if record.get("height") is not None else None,
        duration_sec=float(record["duration_sec"]) if record.get("duration_sec") is not None else None,
        caption=record.get("caption"),
        base64_content=_get_base64_content(media_id) if for_onec and record.get("kind") == "image" else None,
    )


def _message_attachment_payloads(message: Dict[str, object], request: Request, for_onec: bool = False) -> List[AttachmentResponse]:
    return [
        _attachment_response_from_record(item, request, for_onec=for_onec)
        for item in (message.get("attachments") or [])
    ]


def _upload_response_from_media(media, request: Request, for_onec: bool = False) -> UploadResponse:
    attachment = _attachment_response_from_record(
        {
            "id": int(media.media_id),
            "media_id": int(media.media_id),
            "kind": media.kind,
            "mime_type": media.mime_type,
            "size_bytes": int(media.size_bytes),
            "original_name": media.original_name,
            "width": media.width,
            "height": media.height,
            "duration_sec": media.duration_sec,
            "caption": None,
        },
        request,
        for_onec=for_onec,
    )
    return UploadResponse(
        media_id=media.media_id,
        kind=attachment.kind,
        url=attachment.url,
        preview_url=attachment.preview_url,
        mime_type=attachment.mime_type,
        size_bytes=attachment.size_bytes,
        original_name=attachment.original_name,
        width=attachment.width,
        height=attachment.height,
        duration_sec=attachment.duration_sec,
        base64_content=attachment.base64_content,
    )


@router.post("/uploads", response_model=UploadResponse)
async def upload_media(
    request: Request,
    file: UploadFile | None = File(default=None),
    original_name: str | None = Form(default=None),
    mime_type: str | None = Form(default=None),
    _: Dict[str, object] = Depends(get_current_user),
):
    media = await _store_uploaded_media(request, file, original_name, mime_type)
    return _upload_response_from_media(media, request)


@app.post("/integrations/1c/uploads", response_model=UploadResponse)
@router.post("/integrations/1c/uploads", response_model=UploadResponse)
async def upload_onec_media(
    request: Request,
    file: UploadFile | None = File(default=None),
    original_name: str | None = Form(default=None),
    mime_type: str | None = Form(default=None),
    _: None = Depends(require_onec_token),
):
    media = await _store_uploaded_media(request, file, original_name, mime_type)
    return _upload_response_from_media(media, request, for_onec=True)


@app.get("/api/media/{media_id}")
@router.get("/media/{media_id}")
def get_media(
    media_id: int,
    request: Request,
    expires: int | None = Query(default=None),
    signature: str | None = Query(default=None),
    variant: str = Query(default="original"),
):
    media = media_service.get_media_descriptor(media_id)
    if media is None:
        raise HTTPException(status_code=404, detail="Media not found")

    signed_ok = False
    if expires is not None:
        signed_ok = verify_media_access(media_id=media_id, variant=variant, expires_at=expires, signature=signature)
    api_token_header = request.headers.get("X-Api-Token") or request.query_params.get("api_token")
    session_token = request.headers.get("X-Session-Token") or request.query_params.get("session_token")
    has_session = api_token_header == API_TOKEN and session_token and database.get_user_by_session(session_token) is not None
    if not signed_ok and not has_session:
        raise HTTPException(status_code=403, detail="Media access denied")

    direct_url = media_service.get_direct_download_url(media)
    if direct_url:
        return RedirectResponse(direct_url, status_code=307)

    local_path = media_service.get_local_path(media)
    if local_path is None or not local_path.exists():
        raise HTTPException(status_code=404, detail="Stored media file was not found")

    return FileResponse(
        local_path,
        media_type=media.mime_type,
        headers={"Cache-Control": "private, max-age=3600"},
    )


@app.get("/api/media/{media_id}/preview")
@router.get("/media/{media_id}/preview")
def get_media_preview(
    media_id: int,
    request: Request,
    expires: int | None = Query(default=None),
    signature: str | None = Query(default=None),
):
    return get_media(media_id=media_id, request=request, expires=expires, signature=signature, variant="preview")


    return get_media(media_id=media_id, request=request, variant="preview")


@router.get("/chats/{chat_id}/messages", response_model=List[MessageResponse])
def get_chat_messages(
    chat_id: int,
    request: Request,
    limit: int = 50,
    dialog_id: int | None = None,
    current_user: Dict[str, object] = Depends(get_current_user),
):
    if dialog_id is not None:
        dialog = database.get_chat_dialog(dialog_id)
        if dialog is None or dialog["chat_id"] != chat_id:
            raise HTTPException(status_code=404, detail="\u0414\u0438\u0430\u043b\u043e\u0433 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d")
    if not database.user_can_access_chat(
        current_user["id"], current_user["role"], chat_id, dialog_id
    ):
        raise HTTPException(status_code=403, detail="\u041d\u0435\u0442 \u0434\u043e\u0441\u0442\u0443\u043f\u0430 \u043a \u0447\u0430\u0442\u0443")
    allowed_sections = None
    if not database.is_admin_like(current_user["role"]):
        secs = current_user.get("sections")
        if secs:
            allowed_sections = secs
    messages = database.get_messages(
        chat_id,
        limit=limit,
        allowed_sections=allowed_sections,
        dialog_id=dialog_id,
    )
    resolved_dialog_id = dialog_id
    if resolved_dialog_id is None:
        resolved_dialog_id = next(
            (entry.get("dialog_id") for entry in reversed(messages) if entry.get("dialog_id") is not None),
            None,
        )
    if resolved_dialog_id is None:
        active_dialog = database.get_active_chat_dialog(chat_id)
        if active_dialog:
            resolved_dialog_id = int(active_dialog.get("id"))
    result: List[MessageResponse] = []
    for message in messages:
        section_id = message.get("section")
        section_title = None
        if section_id:
            section = next((s for s in database.SECTIONS if s["id"] == section_id), None)
            if section:
                section_title = section["title"]
        stored_message_id = message.get("message_id")
        resolved_message_id = int(stored_message_id) if stored_message_id is not None else int(message["id"])
        result.append(
            MessageResponse(
                id=int(message["id"]),
                message_id=resolved_message_id,
                chat_id=int(message["chat_id"]),
                direction=str(message["direction"]),
                text=str(message["text"]),
                author=message.get("author"),
                created_at=str(message["created_at"]),
                section=section_id,
                section_title=section_title,
                dialog_id=message.get("dialog_id"),
                attachments=_message_attachment_payloads(message, request, for_onec=True),
            )
        )
    if resolved_dialog_id is not None:
        database.mark_dialog_read(current_user["id"], resolved_dialog_id)
    return result


def _resolve_onec_chat_id(external_chat_id: str, explicit_chat_id: int | None) -> int:

    if explicit_chat_id is not None:

        return explicit_chat_id

    digest = hashlib.sha256(external_chat_id.encode("utf-8")).digest()

    hashed = int.from_bytes(digest[:8], "big")

    return ONEC_CHAT_ID_OFFSET + (hashed % ONEC_CHAT_ID_SPACE)





def _normalize_optional(value: str | None) -> str | None:

    if value is None:

        return None

    trimmed = value.strip()

    return trimmed or None





def _parse_int_from_string(value: str | None) -> int | None:

    """Парсит целое число из строки, удаляя пробелы и нечисловые символы."""

    if value is None:

        return None

    # Удаляем все пробелы и нечисловые символы, кроме минуса

    cleaned = re.sub(r'[^\d-]', '', str(value).replace(' ', ''))

    if not cleaned:

        return None

    try:

        return int(cleaned)

    except (ValueError, TypeError):

        return None





@router.post("/messages/send")
def send_message(
    request: ReplyRequest,
    http_request: Request,
    current_user: Dict[str, object] = Depends(get_current_user),
):
    logger.info(
        "Send message: user_id=%s, chat_id=%s, dialog_id=%s, attachments=%s",
        current_user["id"],
        request.chat_id,
        request.dialog_id,
        len(request.attachment_ids),
    )
    if current_user["role"] not in (database.ROLE_ADMIN, database.ROLE_MODERATOR, database.ROLE_OPERATOR):
        raise HTTPException(status_code=403, detail="\u041d\u0435\u0434\u043e\u0441\u0442\u0430\u0442\u043e\u0447\u043d\u043e \u043f\u0440\u0430\u0432 \u0434\u043b\u044f \u043e\u0442\u043f\u0440\u0430\u0432\u043a\u0438 \u0441\u043e\u043e\u0431\u0449\u0435\u043d\u0438\u0439")
    if request.dialog_id is not None:
        dialog = database.get_chat_dialog(request.dialog_id)
        if dialog is None or dialog["chat_id"] != request.chat_id:
            raise HTTPException(status_code=404, detail="\u0414\u0438\u0430\u043b\u043e\u0433 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d")
    if not database.user_can_access_chat(
        current_user["id"], current_user["role"], request.chat_id, request.dialog_id
    ):
        raise HTTPException(status_code=403, detail="\u041d\u0435\u0442 \u0434\u043e\u0441\u0442\u0443\u043f\u0430 \u043a \u0443\u043a\u0430\u0437\u0430\u043d\u043d\u043e\u043c\u0443 \u0447\u0430\u0442\u0443")

    message_text = request.text.strip()
    attachment_ids = request.attachment_ids or []
    if not message_text and not attachment_ids:
        raise HTTPException(status_code=400, detail="Message must contain text or attachments")

    chat = database.get_chat(request.chat_id)
    chat_type = chat.get("type") if chat else None
    section = chat.get("section") if chat else None
    resolved_dialog_id = request.dialog_id or database.get_active_chat_dialog_id(request.chat_id)

    if chat_type == "onec":
        chat_title = chat.get("title") if chat else None
        if not chat_title:
            bin_hint = chat.get("bin") if chat else None
            chat_title = f"1C client {bin_hint}" if bin_hint else f"1C chat {request.chat_id}"
        external_chat_id_value = (chat.get("external_chat_id") if chat else None) or str(request.chat_id)
        try:
            inserted_id = database.save_message(
                chat_id=request.chat_id,
                direction="outgoing",
                text=message_text,
                message_id=None,
                author=current_user["name"],
                chat_title=chat_title,
                username=chat.get("username") if chat else None,
                chat_type=chat_type,
                section=section,
                dialog_id=resolved_dialog_id,
                attachment_ids=attachment_ids,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        if resolved_dialog_id:
            database.set_dialog_operator_mode(resolved_dialog_id, True)

        attachment_records = database.get_message_attachments_map([inserted_id]).get(inserted_id, [])
        attachments_payload = [
            item.dict() for item in [_attachment_response_from_record(record, http_request) for record in attachment_records]
        ]
        _enqueue_onec_outgoing_message(
            message_id=inserted_id,
            chat_id=request.chat_id,
            dialog_id=resolved_dialog_id,
            external_chat_id=external_chat_id_value,
            bin_value=chat.get("bin") if chat else None,
            text=message_text,
            author=current_user["name"],
            section=section,
            attachments=attachments_payload,
        )

        if event_bus.loop:
            asyncio.run_coroutine_threadsafe(
                event_bus.publish_all("new_message", {
                    "chat_id": request.chat_id,
                    "dialog_id": resolved_dialog_id,
                    "message_id": inserted_id,
                    "text": message_text,
                    "direction": "outgoing",
                    "author": current_user["name"],
                    "attachments": attachments_payload,
                }),
                event_bus.loop,
            )

        return {
            "status": "ok",
            "message_id": inserted_id,
            "operator": current_user["name"],
            "dialog_id": resolved_dialog_id,
        }

    sent_message_id: int | None = None
    open_handles = []
    try:
        if attachment_ids:
            media_records = []
            for attachment_id in attachment_ids:
                media = media_service.get_media_descriptor(attachment_id)
                if media is None:
                    raise HTTPException(status_code=400, detail=f"Attachment {attachment_id} was not found")
                media_records.append(media)
            for index, media in enumerate(media_records):
                caption = message_text if index == 0 and message_text else None
                media_payload = media_service.get_direct_download_url(media, expires_in=3600)
                if media_payload is None:
                    local_path = media_service.get_local_path(media)
                    if local_path is None or not local_path.exists():
                        raise HTTPException(status_code=404, detail="Stored media file was not found")
                    handle = local_path.open("rb")
                    open_handles.append(handle)
                    media_payload = handle
                if media.kind == "image":
                    sent = bot.send_photo(request.chat_id, media_payload, caption=caption)
                else:
                    sent = bot.send_video(request.chat_id, media_payload, caption=caption, supports_streaming=True)
                if sent_message_id is None:
                    sent_message_id = sent.message_id
        elif message_text:
            sent_message_id = bot.send_message(request.chat_id, message_text).message_id
    except HTTPException:
        raise
    except Exception as exc:  # pragma: no cover - depends on Telegram API
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    finally:
        for handle in open_handles:
            try:
                handle.close()
            except Exception:
                pass

    try:
        stored_message_id = database.save_message(
            chat_id=request.chat_id,
            direction="outgoing",
            text=message_text,
            message_id=sent_message_id,
            author=current_user["name"],
            chat_title=(chat.get("title") if chat else None) or str(request.chat_id),
            username=chat.get("username") if chat else None,
            chat_type=chat_type or "private",
            section=section,
            dialog_id=resolved_dialog_id,
            attachment_ids=attachment_ids,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    attachment_records = database.get_message_attachments_map([stored_message_id]).get(stored_message_id, [])
    attachments_payload = [
        item.dict() for item in [_attachment_response_from_record(record, http_request) for record in attachment_records]
    ]

    if event_bus.loop:
        asyncio.run_coroutine_threadsafe(
            event_bus.publish_all("new_message", {
                "chat_id": request.chat_id,
                "dialog_id": resolved_dialog_id,
                "message_id": stored_message_id,
                "text": message_text,
                "direction": "outgoing",
                "author": current_user["name"],
                "attachments": attachments_payload,
            }),
            event_bus.loop,
        )

    return {
        "status": "ok",
        "message_id": sent_message_id or stored_message_id,
        "operator": current_user["name"],
        "dialog_id": resolved_dialog_id,
    }


@router.post("/dialogs/{dialog_id}/ai/enable")

def enable_ai_for_dialog(

    dialog_id: int, current_user: Dict[str, object] = Depends(get_current_user)

):

    dialog = database.get_chat_dialog(dialog_id)

    if dialog is None:

        raise HTTPException(status_code=404, detail="Диалог не найден")



    chat_id = int(dialog["chat_id"])

    if not database.user_can_access_chat(

        current_user["id"], current_user["role"], chat_id, dialog_id

    ):

        raise HTTPException(status_code=403, detail="Нет доступа к диалогу")



    chat = database.get_chat(chat_id)

    if chat is None:

        raise HTTPException(status_code=404, detail="Чат не найден")



    database.set_dialog_operator_mode(dialog_id, False)



    section = chat.get("section") if chat else None

    chat_title = chat.get("title") if chat else None

    chat_type = chat.get("type") if chat else None

    notification_text = (

        "🤖 AI помощник снова включён для этого диалога. Можно продолжать задавать вопросы."

    )



    if chat_type == "onec":

        message_id = database.save_message(

            chat_id=chat_id,

            direction="outgoing",

            text=notification_text,

            message_id=None,

            author="System",

            chat_title=chat_title or str(chat_id),

            username=None,

            chat_type=chat_type,

            section=section,

            dialog_id=dialog_id,

        )

        external_chat_id = chat.get("external_chat_id") or str(chat_id)

        _enqueue_onec_outgoing_message(

            message_id=message_id,

            chat_id=chat_id,

            dialog_id=dialog_id,

            external_chat_id=external_chat_id,

            bin_value=chat.get("bin"),

            text=notification_text,

            author="System",

            section=section,

        )

        return {

            "status": "ok",

            "chat_id": chat_id,

            "dialog_id": dialog_id,

            "message_id": message_id,

        }



    try:

        sent_message = bot.send_message(chat_id, notification_text)

    except Exception as exc:  # pragma: no cover - depends on Telegram API

        raise HTTPException(status_code=502, detail=str(exc)) from exc



    enable_ai_session(chat_id)



    database.save_message(

        chat_id=chat_id,

        direction="outgoing",

        text=notification_text,

        message_id=sent_message.message_id,

        author="System",

        chat_title=chat_title or sent_message.chat.title or sent_message.chat.username or str(chat_id),

        username=sent_message.chat.username,

        chat_type=sent_message.chat.type,

        section=section,

        dialog_id=dialog_id,

    )

    return {

        "status": "ok",

        "chat_id": chat_id,

        "dialog_id": dialog_id,

        "message_id": sent_message.message_id,

    }





@router.post("/dialogs/{dialog_id}/ai/disable")

def disable_ai_for_dialog(

    dialog_id: int, current_user: Dict[str, object] = Depends(get_current_user)

):

    dialog = database.get_chat_dialog(dialog_id)

    if dialog is None:

        raise HTTPException(status_code=404, detail="Диалог не найден")



    chat_id = int(dialog["chat_id"])

    if not database.user_can_access_chat(

        current_user["id"], current_user["role"], chat_id, dialog_id

    ):

        raise HTTPException(status_code=403, detail="Нет доступа к диалогу")



    chat = database.get_chat(chat_id)

    if chat is None:

        raise HTTPException(status_code=404, detail="Чат не найден")



    database.set_dialog_operator_mode(dialog_id, True)



    section = chat.get("section") if chat else None

    chat_title = chat.get("title") if chat else None

    chat_type = chat.get("type") if chat else None

    notification_text = (

        "👨‍💼 Оператор подключён. AI помощник отключён для этого диалога."

    )



    if chat_type == "onec":

        message_id = database.save_message(

            chat_id=chat_id,

            direction="outgoing",

            text=notification_text,

            message_id=None,

            author="System",

            chat_title=chat_title or str(chat_id),

            username=None,

            chat_type=chat_type,

            section=section,

            dialog_id=dialog_id,

        )

        external_chat_id = chat.get("external_chat_id") or str(chat_id)

        _enqueue_onec_outgoing_message(

            message_id=message_id,

            chat_id=chat_id,

            dialog_id=dialog_id,

            external_chat_id=external_chat_id,

            bin_value=chat.get("bin"),

            text=notification_text,

            author="System",

            section=section,

        )

        return {

            "status": "ok",

            "chat_id": chat_id,

            "dialog_id": dialog_id,

            "message_id": message_id,

        }



    try:

        sent_message = bot.send_message(chat_id, notification_text)

    except Exception as exc:  # pragma: no cover - depends on Telegram API

        raise HTTPException(status_code=502, detail=str(exc)) from exc



    database.save_message(

        chat_id=chat_id,

        direction="outgoing",

        text=notification_text,

        message_id=sent_message.message_id,

        author="System",

        chat_title=chat_title

        or sent_message.chat.title

        or sent_message.chat.username

        or str(chat_id),

        username=sent_message.chat.username,

        chat_type=sent_message.chat.type,

        section=section,

        dialog_id=dialog_id,

    )



    return {

        "status": "ok",

        "chat_id": chat_id,

        "dialog_id": dialog_id,

        "message_id": sent_message.message_id,

    }





@router.post("/dialogs/{dialog_id}/close", response_model=DialogStatusResponse)

def close_dialog(

    dialog_id: int, current_user: Dict[str, object] = Depends(get_current_user)

):

    dialog = database.get_chat_dialog(dialog_id)

    if dialog is None:

        raise HTTPException(status_code=404, detail="Диалог не найден")



    chat_id = int(dialog["chat_id"])

    if not database.user_can_access_chat(

        current_user["id"], current_user["role"], chat_id, dialog_id

    ):

        raise HTTPException(status_code=403, detail="Нет доступа к диалогу")



    chat = database.get_chat(chat_id)

    if chat is None:

        raise HTTPException(status_code=404, detail="Чат не найден")



    if database.close_chat_dialog(dialog_id) is None:

        raise HTTPException(status_code=404, detail="Диалог не найден")



    database.set_dialog_operator_mode(dialog_id, False)



    section = chat.get("section") if chat else None

    chat_title = chat.get("title") if chat else None

    chat_type = chat.get("type") if chat else None

    notification_text = (

        "Обращение закрыто оператором. "

        "🤖 AI снова включён. Напишите новое сообщение, чтобы возобновить диалог."

    )



    closed_at = datetime.now(timezone.utc).isoformat()
    latest_appeal_id = database.get_latest_closed_appeal_id(dialog_id)
    employee_assessments = database.create_employee_client_assessments_for_dialog(
        dialog_id,
        appeal_id=latest_appeal_id,
    )
    employee_assessment_id = (
        int(employee_assessments[0]["id"])
        if employee_assessments and employee_assessments[0].get("id") is not None
        else None
    )
    if chat_type == "onec":

        message_id = database.save_message(

            chat_id=chat_id,

            direction="outgoing",

            text=notification_text,

            message_id=None,

            author="System",

            chat_title=chat_title or str(chat_id),

            username=None,

            chat_type=chat_type,

            section=section,

            dialog_id=dialog_id,

        )

        external_chat_id = chat.get("external_chat_id") or str(chat_id)

        _enqueue_onec_outgoing_message(

            message_id=message_id,

            chat_id=chat_id,

            dialog_id=dialog_id,

            external_chat_id=external_chat_id,

            bin_value=chat.get("bin"),

            text=notification_text,

            author="System",

            section=section,

        )
        return DialogStatusResponse(

            chat_id=chat_id,

            dialog_id=dialog_id,

            dialog_closed_at=closed_at,

            dialog_purge_at=database.calculate_dialog_purge_at(closed_at),

            ai_enabled=True,
            employee_assessment_id=employee_assessment_id,
            employee_assessment_pending=employee_assessment_id is not None,

        )



    try:

        sent_message = bot.send_message(chat_id, notification_text)

    except Exception as exc:  # pragma: no cover - depends on Telegram API

        raise HTTPException(status_code=502, detail=str(exc)) from exc



    enable_ai_session(chat_id)



    database.save_message(

        chat_id=chat_id,

        direction="outgoing",

        text=notification_text,

        message_id=sent_message.message_id,

        author="System",

        chat_title=chat_title

        or sent_message.chat.title

        or sent_message.chat.username

        or str(chat_id),

        username=sent_message.chat.username,

        chat_type=sent_message.chat.type,

        section=section,

        dialog_id=dialog_id,

    )



    # Send CSAT rating request to the client

    try:

        send_csat_request(chat_id, dialog_id, latest_appeal_id)

    except Exception:

        logger.warning("Failed to send CSAT request for dialog %s", dialog_id, exc_info=True)

    try:
        survey_service.maybe_start_survey_after_appeal_closed(dialog_id, latest_appeal_id)
    except Exception:
        logger.warning("Failed to start after-close survey for dialog %s", dialog_id, exc_info=True)



    return DialogStatusResponse(

        chat_id=chat_id,

        dialog_id=dialog_id,

        dialog_closed_at=closed_at,

        dialog_purge_at=database.calculate_dialog_purge_at(closed_at),

        ai_enabled=True,
        employee_assessment_id=employee_assessment_id,
        employee_assessment_pending=employee_assessment_id is not None,

    )





@router.post("/dialogs/{dialog_id}/open", response_model=DialogStatusResponse)

def open_dialog(

    dialog_id: int, current_user: Dict[str, object] = Depends(get_current_user)

):

    dialog = database.get_chat_dialog(dialog_id)

    if dialog is None:

        raise HTTPException(status_code=404, detail="Диалог не найден")



    chat_id = int(dialog["chat_id"])

    if not database.user_can_access_chat(

        current_user["id"], current_user["role"], chat_id, dialog_id

    ):

        raise HTTPException(status_code=403, detail="Нет доступа к диалогу")



    activated = database.activate_chat_dialog(dialog_id, chat_id=chat_id)

    if activated is None:

        raise HTTPException(status_code=404, detail="Диалог не найден")



    database.set_dialog_operator_mode(dialog_id, False)



    return DialogStatusResponse(

        chat_id=chat_id,

        dialog_id=dialog_id,

        dialog_closed_at=None,

        dialog_purge_at=None,

        ai_enabled=not database.is_dialog_in_operator_mode(dialog_id),

    )



def _publish_new_message_event(
    *,
    chat_id: int,
    dialog_id: int | None,
    message_id: int,
    text: str,
    direction: Literal["incoming", "outgoing"],
    author: str | None,
    attachments: list[dict] | None = None,
) -> None:
    if event_bus.loop:
        asyncio.run_coroutine_threadsafe(
            event_bus.publish_all(
                "new_message",
                {
                    "chat_id": chat_id,
                    "dialog_id": dialog_id,
                    "message_id": message_id,
                    "text": text,
                    "direction": direction,
                    "author": author,
                    "attachments": attachments or [],
                },
            ),
            event_bus.loop,
        )


def _store_onec_outgoing_text_message(
    *,
    chat_id: int,
    dialog_id: int | None,
    external_chat_id: str,
    bin_value: str | None,
    text: str,
    author: str | None,
    chat_title: str,
    section: str | None,
    quick_replies: List[dict] | None = None,
) -> int:
    message_id = database.save_message(
        chat_id=chat_id,
        direction="outgoing",
        text=text,
        message_id=None,
        author=author,
        chat_title=chat_title,
        username=None,
        chat_type="onec",
        section=section,
        dialog_id=dialog_id,
        quick_replies=quick_replies,
    )
    _enqueue_onec_outgoing_message(
        message_id=message_id,
        chat_id=chat_id,
        dialog_id=dialog_id,
        external_chat_id=external_chat_id,
        bin_value=bin_value,
        text=text,
        author=author,
        section=section,
        quick_replies=quick_replies,
    )
    _publish_new_message_event(
        chat_id=chat_id,
        dialog_id=dialog_id,
        message_id=message_id,
        text=text,
        direction="outgoing",
        author=author,
    )
    return message_id


def _process_onec_incoming_message(
    *,
    chat_id: int,
    dialog_id: int,
    external_chat_id: str,
    bin_value: str,
    message_text: str,
    normalized_text: str,
    author: str | None,
    chat_title: str,
    section_id: str | None,
) -> None:
    try:
        chat_record = database.get_chat(chat_id)
        chat_section = section_id or (chat_record.get("section") if chat_record else None)
        chat_bin = (chat_record.get("bin") if chat_record else None) or bin_value

        language_selection = _normalize_onec_language_command(normalized_text)
        if language_selection is not None:
            _, language_notice = language_selection
            _store_onec_outgoing_text_message(
                chat_id=chat_id,
                dialog_id=dialog_id,
                external_chat_id=external_chat_id,
                bin_value=chat_bin,
                text=language_notice,
                author="System",
                chat_title=chat_title,
                section=chat_section,
                quick_replies=_onec_default_quick_replies(),
            )
            return

        if _is_onec_operator_request(normalized_text):
            database.set_dialog_operator_mode(dialog_id, True)
            operator_notice = "Ваш запрос передан оператору. AI помощник отключён до закрытия диалога."
            _store_onec_outgoing_text_message(
                chat_id=chat_id,
                dialog_id=dialog_id,
                external_chat_id=external_chat_id,
                bin_value=chat_bin,
                text=operator_notice,
                author="System",
                chat_title=chat_title,
                section=chat_section,
            )
            database.create_operator_request_notifications(
                chat_id,
                dialog_id=dialog_id,
                chat_title=chat_title,
                section=chat_section,
                bin_value=chat_bin,
            )
            return

        if survey_service.handle_channel_survey_text_answer(chat_id, message_text):
            return

        if database.is_dialog_in_operator_mode(dialog_id):
            return

        faq_entry = database.find_faq_entry_by_keywords(message_text, chat_section)
        if faq_entry:
            response_section = faq_entry.get("section") or chat_section
            response_text = faq_entry.get("answer", "").strip()
            response_question = faq_entry.get("question", "").strip()
            if response_question:
                response_text = f"FAQ:\\n{response_question}\\n\\n{response_text}" if response_text else response_question
            if not response_text:
                response_text = "Готовый ответ не найден. При необходимости нажмите «Позвать оператора»."
            if response_section and response_section != chat_section:
                database.set_chat_section(chat_id, response_section, dialog_id=dialog_id)
                chat_section = response_section
            database.set_dialog_operator_mode(dialog_id, False)
            _store_onec_outgoing_text_message(
                chat_id=chat_id,
                dialog_id=dialog_id,
                external_chat_id=external_chat_id,
                bin_value=chat_bin,
                text=response_text,
                author="AutoBot",
                chat_title=chat_title,
                section=chat_section,
                quick_replies=_onec_default_quick_replies(),
            )
            return

        history = database.get_messages(chat_id, limit=6, dialog_id=dialog_id)
        if ai_manager is not None:
            ai_reply = ai_manager.generate_response(message_text, history)
        else:
            ai_reply = (
                "AI помощник временно недоступен. При необходимости нажмите «Позвать оператора»."
            )

        ai_reply = (ai_reply or "").strip()
        if not ai_reply:
            return

        database.set_dialog_operator_mode(dialog_id, False)
        _store_onec_outgoing_text_message(
            chat_id=chat_id,
            dialog_id=dialog_id,
            external_chat_id=external_chat_id,
            bin_value=chat_bin,
            text=ai_reply,
            author="AI Assistant",
            chat_title=chat_title,
            section=chat_section,
            quick_replies=_onec_default_quick_replies(),
        )
    except Exception:
        logger.exception(
            "Failed to post-process 1C message: ext_id=%s chat_id=%s dialog_id=%s",
            external_chat_id,
            chat_id,
            dialog_id,
        )


@app.post("/integrations/1c/messages")
@router.post("/integrations/1c/messages")
def create_onec_message(
    request: OneCIncomingMessageRequest,
    http_request: Request,
    background_tasks: BackgroundTasks,
    _: None = Depends(require_onec_token),
):
    section_id = _normalize_optional(request.section)
    if section_id and not any(section["id"] == section_id for section in database.SECTIONS):
        section_id = None

    bin_value = _normalize_optional(request.bin)
    if not bin_value:
        raise HTTPException(status_code=400, detail="BIN is required")

    external_chat_id = request.external_chat_id.strip()
    if not external_chat_id:
        raise HTTPException(status_code=400, detail="external_chat_id is required")

    message_text = request.text.strip()
    chat_id = _resolve_onec_chat_id(external_chat_id, request.chat_id)
    author = _normalize_optional(request.author)
    normalized_text = message_text.lower()
    language_selection = _normalize_onec_language_command(normalized_text)
    active_survey_session = None
    survey_dialog_id: int | None = None
    if _is_onec_operator_request(normalized_text):
        stored_text = "[OPERATOR REQUEST]"
    elif language_selection is not None:
        stored_text = f"[LANGUAGE:{language_selection[0]}]"
    else:
        active_survey_session = database.get_active_survey_session(int(chat_id))
        if active_survey_session and active_survey_session.get("dialog_id") is not None:
            survey_dialog_id = int(active_survey_session["dialog_id"])
        stored_text = survey_service.get_channel_survey_answer_display_text(chat_id, message_text)
        message_text = stored_text
        normalized_text = message_text.lower()

    title_candidate = _normalize_optional(request.title)
    chat_title = title_candidate or f"1C client {external_chat_id}"

    database.upsert_chat(chat_id, chat_title, None, "onec", external_chat_id=external_chat_id)
    if survey_dialog_id is not None:
        dialog_id = survey_dialog_id
        dialog_resumed = False
        is_survey_answer = True
    else:
        try:
            dialog_result = database.ensure_active_chat_dialog(
                chat_id,
                bin_value,
                section=section_id,
                return_state=True,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=500, detail=str(exc)) from exc
        if isinstance(dialog_result, tuple):
            dialog_id = int(dialog_result[0])
            dialog_resumed = bool(dialog_result[1])
        else:
            dialog_id = int(dialog_result)
            dialog_resumed = False
        is_survey_answer = False

    try:
        message_id = database.save_message(
            chat_id=chat_id,
            direction="incoming",
            text=stored_text,
            message_id=None,
            author=author,
            chat_title=chat_title,
            username=None,
            chat_type="onec",
            section=section_id,
            dialog_id=dialog_id,
            attachment_ids=request.attachment_ids,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    logger.info(
        "1C incoming message saved: ext_id=%s, chat_id=%s, dialog_id=%s, message_id=%s, author=%s",
        external_chat_id, chat_id, dialog_id, message_id, author,
    )

    incoming_attachment_records = database.get_message_attachments_map([message_id]).get(message_id, [])
    incoming_attachments_payload = [
        item.dict() for item in [_attachment_response_from_record(record, http_request) for record in incoming_attachment_records]
    ]
    _publish_new_message_event(
        chat_id=chat_id,
        dialog_id=dialog_id,
        message_id=message_id,
        text=stored_text,
        direction="incoming",
        author=author,
        attachments=incoming_attachments_payload,
    )

    if section_id and not is_survey_answer:
        database.set_chat_section(chat_id, section_id, dialog_id=dialog_id)

    if dialog_resumed:
        _store_onec_outgoing_text_message(
            chat_id=chat_id,
            dialog_id=dialog_id,
            external_chat_id=external_chat_id,
            bin_value=bin_value,
            text="Диалог возобновлён. Новое обращение открыто, AI снова включён.",
            author="System",
            chat_title=chat_title,
            section=section_id,
            quick_replies=_onec_default_quick_replies(),
        )

    if is_survey_answer:
        is_first_message_in_dialog = False
        has_contract = True
    else:
        existing_messages = database.get_messages(chat_id, limit=2, dialog_id=dialog_id)
        is_first_message_in_dialog = len(existing_messages) <= 1
        has_contract = False if is_first_message_in_dialog else not database.has_organization_without_contract(bin_value)
    response_message = None

    if is_first_message_in_dialog:
        logger.info("Checking contract for 1C customer BIN: %s", bin_value)
        contract_result = contract_checker.check_customer_contracts(bin_value)
        has_contract = _persist_bin_contract_result(bin_value, contract_result)
        logger.info("Contract check result for %s: has_contract=%s", bin_value, has_contract)
        response_message = _build_onec_contract_status_text(
            has_contract=bool(has_contract),
            year=contract_checker.ACTIVE_CONTRACT_YEAR,
        )
        _store_onec_outgoing_text_message(
            chat_id=chat_id,
            dialog_id=dialog_id,
            external_chat_id=external_chat_id,
            bin_value=bin_value,
            text=response_message,
            author="System",
            chat_title=chat_title,
            section=section_id,
            quick_replies=_onec_language_quick_replies(),
        )

    if not message_text:
        return {
            "status": "ok",
            "has_contract": has_contract,
            "response_message": response_message,
            "chat_id": chat_id,
            "dialog_id": dialog_id,
        }

    background_tasks.add_task(
        _process_onec_incoming_message,
        chat_id=chat_id,
        dialog_id=dialog_id,
        external_chat_id=external_chat_id,
        bin_value=bin_value,
        message_text=message_text,
        normalized_text=normalized_text,
        author=author,
        chat_title=chat_title,
        section_id=section_id,
    )

    return {
        "status": "ok",
        "has_contract": has_contract,
        "response_message": response_message,
        "chat_id": chat_id,
        "dialog_id": dialog_id,
    }


# ------------------ 1C history ------------------

def _onec_history_core(
    external_chat_id: str,
    chat_id: int | None,
    dialog_id: int | None,
    bin_value: str | None,
    limit: int,
    request: Request,
) -> OneCMessagesResponse:
    normalized_external = external_chat_id.strip()
    if not normalized_external:
        raise HTTPException(status_code=400, detail="external_chat_id is required")

    resolved_chat_id = _resolve_onec_chat_id(normalized_external, chat_id)

    if dialog_id is not None:
        dialog = database.get_chat_dialog(dialog_id)
        if dialog is None or dialog["chat_id"] != resolved_chat_id:
            raise HTTPException(status_code=404, detail="\u0414\u0438\u0430\u043b\u043e\u0433 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d")
        resolved_dialog_id = dialog_id
    else:
        normalized_bin = (bin_value or "").strip() or None
        if normalized_bin is not None:
            database.upsert_chat(resolved_chat_id, "External User", None, "onec", external_chat_id=normalized_external)
            resolved_dialog_id = None
            active_dialog = database.get_active_chat_dialog(resolved_chat_id)
            if active_dialog and str(active_dialog.get("bin") or "").strip() == normalized_bin:
                resolved_dialog_id = int(active_dialog["id"])
            else:
                latest_closed_dialog_id = database.get_latest_closed_chat_dialog_id(resolved_chat_id)
                if latest_closed_dialog_id is not None:
                    latest_closed_dialog = database.get_chat_dialog(latest_closed_dialog_id)
                    if latest_closed_dialog and str(latest_closed_dialog.get("bin") or "").strip() == normalized_bin:
                        resolved_dialog_id = int(latest_closed_dialog["id"])
        else:
            resolved_dialog_id = None

    chat = database.get_chat(resolved_chat_id)
    if chat and chat.get("type") not in (None, "onec"):
        raise HTTPException(status_code=403, detail="\u0427\u0430\u0442 \u043d\u0435 \u043f\u0440\u0435\u0434\u043d\u0430\u0437\u043d\u0430\u0447\u0435\u043d \u0434\u043b\u044f \u0438\u043d\u0442\u0435\u0433\u0440\u0430\u0446\u0438\u0438 1\u0421")

    raw_messages = database.get_messages(resolved_chat_id, limit=limit, dialog_id=resolved_dialog_id)
    logger.info(
        "1C history request: ext_id=%s, chat_id=%s, dialog_id=%s, messages_count=%d",
        normalized_external, resolved_chat_id, resolved_dialog_id, len(raw_messages),
    )

    messages: List[OneCMessageEntry] = []
    for message in reversed(raw_messages):
        stored_message_id = message.get("message_id")
        if stored_message_id is None:
            stored_message_id = message.get("id")
        created_at_value = message.get("created_at")
        created_at_iso = created_at_value.isoformat() if isinstance(created_at_value, datetime) else str(created_at_value)
        messages.append(
            OneCMessageEntry(
                message_id=int(stored_message_id) if stored_message_id is not None else None,
                chat_id=int(message["chat_id"]),
                dialog_id=message.get("dialog_id"),
                direction=str(message["direction"]),
                text=str(message["text"]),
                author=message.get("author"),
                created_at=created_at_iso,
                section=message.get("section"),
                attachments=_message_attachment_payloads(message, request, for_onec=True),
                quick_replies=message.get("quick_replies") or [],
            )
        )

    if resolved_dialog_id is None:
        resolved_dialog_id = database.get_active_chat_dialog_id(resolved_chat_id)

    return OneCMessagesResponse(
        external_chat_id=normalized_external,
        chat_id=resolved_chat_id,
        dialog_id=resolved_dialog_id,
        messages=messages,
    )


@app.get(
    "/integrations/1c/messages",
    response_model=OneCMessagesResponse,
)
@router.get(
    "/integrations/1c/messages",
    response_model=OneCMessagesResponse,
)
def list_onec_messages(
    request: Request,
    external_chat_id: str = Query(min_length=1, max_length=128),
    _: None = Depends(require_onec_token),
    chat_id: str | None = Query(default=None),
    dialog_id: str | None = Query(default=None),
    bin: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=500),
):
    parsed_chat_id = _parse_int_from_string(chat_id) if chat_id else None
    parsed_dialog_id = _parse_int_from_string(dialog_id) if dialog_id else None
    return _onec_history_core(external_chat_id, parsed_chat_id, parsed_dialog_id, bin, limit, request)


class OneCHistoryPostRequest(BaseModel):
    external_chat_id: str = Field(min_length=1, max_length=128)
    chat_id: str | None = None
    dialog_id: str | None = None
    bin: str | None = None
    limit: int = Field(default=200, ge=1, le=500)


@app.post("/integrations/1c/messages/history", response_model=OneCMessagesResponse)
@router.post("/integrations/1c/messages/history", response_model=OneCMessagesResponse)
def list_onec_messages_post(
    body: OneCHistoryPostRequest,
    request: Request,
    _: None = Depends(require_onec_token),
):
    parsed_chat_id = _parse_int_from_string(body.chat_id) if body.chat_id else None
    parsed_dialog_id = _parse_int_from_string(body.dialog_id) if body.dialog_id else None
    return _onec_history_core(body.external_chat_id, parsed_chat_id, parsed_dialog_id, body.bin, body.limit, request)


# ------------------ Outbox 1? ------------------

class OneCOutboxItem(BaseModel):

    outbox_id: int

    payload: dict

    signature: str | None = None





class OneCOutboxResponse(BaseModel):

    items: List[OneCOutboxItem] = Field(default_factory=list)





class OneCAckRequest(BaseModel):

    delivered_ids: List[int] = Field(default_factory=list)

    failed_ids: List[Dict[str, object]] = Field(default_factory=list)  # [{"id": 1, "error": "text"}]





@app.get("/integrations/1c/outbox", response_model=OneCOutboxResponse)

@router.get("/integrations/1c/outbox", response_model=OneCOutboxResponse)

def onec_outbox(

    external_chat_id: str = Query(min_length=1, max_length=128),

    limit: int = Query(default=100, ge=1, le=500),

    _: None = Depends(require_onec_token),

):

    """1C pulls operator messages (Web->Backend) for the specified external_chat_id."""

    items = []

    rows = database.outbox_list_pending_onec(external_chat_id, limit)

    for r in rows:

        payload = r["payload"] or {}

        sig = payload.get("signature") or _sign_payload(payload)

        items.append(OneCOutboxItem(outbox_id=r["id"], payload=payload, signature=sig or None))

    return OneCOutboxResponse(items=items)





@app.post("/integrations/1c/ack")

@router.post("/integrations/1c/ack")

def onec_ack(

    body: OneCAckRequest,

    _: None = Depends(require_onec_token),

):

    """1С подтверждает доставку (delivered_ids) и/или сообщает failed_ids с ошибкой."""

    try:

        if body.delivered_ids:

            database.outbox_mark_delivered_onec(body.delivered_ids)

        if body.failed_ids:

            database.outbox_mark_failed_onec(body.failed_ids)

    except Exception as exc:

        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return {"status": "ok"}





class OneCCloseRequest(BaseModel):

    external_chat_id: str = Field(min_length=1, max_length=128)

    chat_id: int | None = None

    bin: str | None = None





class OneCRatingRequest(BaseModel):

    external_chat_id: str = Field(min_length=1, max_length=128)

    chat_id: int | None = None

    dialog_id: int | None = None

    appeal_id: int | None = None

    rating: int = Field(ge=1, le=5)

    target: Literal["operator", "ai"] = "operator"


class SurveyTemplateRequest(BaseModel):
    title: str
    description: str = ""
    audience: str = "client"
    status: str = "draft"
    trigger_type: str = "periodic"
    periodic_interval: str | None = None
    scheduled_at: str | None = None
    launch_rules: List[Dict[str, Any]] = Field(default_factory=list)
    is_anonymous: bool = False
    questions: List[Dict[str, Any]] = Field(default_factory=list)


class ManualSurveyLaunchRequest(BaseModel):
    template_id: int
    bin_values: List[str] = Field(default_factory=list)
    dialog_ids: List[int] = Field(default_factory=list)


class EmployeeClientAssessmentSubmitRequest(BaseModel):
    question_clarity_score: int = Field(ge=1, le=5)
    data_completeness_score: int = Field(ge=1, le=5)
    client_response_speed_score: int = Field(ge=1, le=5)
    business_communication_score: int = Field(ge=1, le=5)
    client_readiness_score: int = Field(ge=1, le=5)
    low_score_reason: str | None = None
    internal_comment: str | None = None
    interaction_status: str
    interaction_flag: str
    request_repeat_status: str
    client_data_overdue: bool = False


def _parse_optional_date_param(value: str | None, *, field_name: str) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"Некорректная дата: {field_name}") from exc

@app.post("/integrations/1c/close")

@router.post("/integrations/1c/close")

def onec_close_dialog(

    body: OneCCloseRequest,

    _: None = Depends(require_onec_token),

):

    """1С закрывает активный диалог (обращение) клиента."""

    external_chat_id = body.external_chat_id.strip()

    if not external_chat_id:

        raise HTTPException(status_code=400, detail="external_chat_id \u043e\u0431\u044f\u0437\u0430\u0442\u0435\u043b\u0435\u043d")



    chat_id = _resolve_onec_chat_id(external_chat_id, body.chat_id)

    active = database.get_active_chat_dialog(chat_id)

    if active is None:

        return {"status": "no_active_dialog", "message": "\u041d\u0435\u0442 \u0430\u043a\u0442\u0438\u0432\u043d\u043e\u0433\u043e \u0434\u0438\u0430\u043b\u043e\u0433\u0430 \u0434\u043b\u044f \u0437\u0430\u043a\u0440\u044b\u0442\u0438\u044f"}



    dialog_id = int(active["id"])

    database.close_chat_dialog(dialog_id, closed_by="client")

    latest_stats = database.get_latest_dialog_stats(dialog_id)

    latest_appeal_id = (

        int(latest_stats["appeal_id"])

        if latest_stats and latest_stats.get("appeal_id") is not None

        else database.get_latest_closed_appeal_id(dialog_id)

    )

    rating_target = "ai" if latest_stats and latest_stats.get("is_ai_closed") else "operator"

    rating_required = latest_appeal_id is not None
    employee_assessments = database.create_employee_client_assessments_for_dialog(
        dialog_id,
        appeal_id=latest_appeal_id,
    )
    employee_assessment_id = (
        int(employee_assessments[0]["id"])
        if employee_assessments and employee_assessments[0].get("id") is not None
        else None
    )



    # Notify Telegram client if applicable

    try:

        from backend.telegram_bot import bot as tg_bot

        tg_bot.send_message(

            chat_id,

            "\u041e\u0431\u0440\u0430\u0449\u0435\u043d\u0438\u0435 \u0437\u0430\u0432\u0435\u0440\u0448\u0435\u043d\u043e \u043a\u043b\u0438\u0435\u043d\u0442\u043e\u043c \u0438\u0437 1\u0421. AI \u0441\u043d\u043e\u0432\u0430 \u0432\u043a\u043b\u044e\u0447\u0451\u043d.\n"

            "\u041d\u0430\u043f\u0438\u0448\u0438\u0442\u0435 \u043d\u043e\u0432\u043e\u0435 \u0441\u043e\u043e\u0431\u0449\u0435\u043d\u0438\u0435, \u0447\u0442\u043e\u0431\u044b \u0432\u043e\u0437\u043e\u0431\u043d\u043e\u0432\u0438\u0442\u044c \u0434\u0438\u0430\u043b\u043e\u0433."

        )

        if rating_required and latest_appeal_id is not None:

            if rating_target == "ai":

                send_ai_csat_request(chat_id, dialog_id, latest_appeal_id)

            else:

                send_csat_request(chat_id, dialog_id, latest_appeal_id)

    except Exception:

        pass  # Telegram notification is best-effort

    logger.info(

        "1C client closed dialog: ext_id=%s, chat_id=%s, dialog_id=%s",

        external_chat_id, chat_id, dialog_id,

    )

    return {

        "status": "ok",

        "dialog_id": dialog_id,

        "appeal_id": latest_appeal_id,

        "rating_required": bool(rating_required),

        "rating_target": rating_target,
        "employee_assessment_count": len(employee_assessments),
        "employee_assessment_id": employee_assessment_id,
        "employee_assessment_pending": employee_assessment_id is not None,

    }





@app.post("/integrations/1c/rating")

@router.post("/integrations/1c/rating")

def onec_submit_rating(

    body: OneCRatingRequest,

    _: None = Depends(require_onec_token),

):

    """1С сохраняет оценку качества обслуживания (оператор или AI)."""

    external_chat_id = body.external_chat_id.strip()

    if not external_chat_id:

        raise HTTPException(status_code=400, detail="external_chat_id \u043e\u0431\u044f\u0437\u0430\u0442\u0435\u043b\u0435\u043d")



    chat_id = _resolve_onec_chat_id(external_chat_id, body.chat_id)

    appeal_id = body.appeal_id

    dialog_id = body.dialog_id

    if appeal_id is not None and dialog_id is None:

        dialog_id = database.get_dialog_id_for_appeal(appeal_id)

    if dialog_id is None:

        dialog_id = database.get_latest_closed_chat_dialog_id(chat_id)

    if dialog_id is None and body.chat_id is None:
        stored_chat = database.get_chat_by_external_chat_id(external_chat_id)
        if stored_chat and stored_chat.get("chat_id") is not None:
            stored_chat_id = int(stored_chat["chat_id"])
            if stored_chat_id != chat_id:
                chat_id = stored_chat_id
                dialog_id = database.get_latest_closed_chat_dialog_id(chat_id)



    if dialog_id is None:

        logger.warning(
            "1C rating rejected: no closed dialog found ext_id=%s resolved_chat_id=%s explicit_chat_id=%s appeal_id=%s target=%s",
            external_chat_id,
            chat_id,
            body.chat_id,
            appeal_id,
            body.target,
        )
        raise HTTPException(status_code=400, detail="dialog_id \u0438\u043b\u0438 appeal_id \u043e\u0431\u044f\u0437\u0430\u0442\u0435\u043b\u0435\u043d")



    if appeal_id is None:

        appeal_id = database.get_latest_closed_appeal_id(dialog_id)



    if body.target == "ai":

        saved = database.save_ai_csat_rating(
            dialog_id,
            body.rating,
            appeal_id=appeal_id,
            rater_external_chat_id=external_chat_id,
            channel=database.RATING_CHANNEL_ONEC_API,
        )

    else:

        saved = database.save_csat_rating(
            dialog_id,
            body.rating,
            appeal_id=appeal_id,
            rater_external_chat_id=external_chat_id,
            channel=database.RATING_CHANNEL_ONEC_API,
        )



    if not saved:

        raise HTTPException(status_code=404, detail="\u041d\u0435 \u0443\u0434\u0430\u043b\u043e\u0441\u044c \u0441\u043e\u0445\u0440\u0430\u043d\u0438\u0442\u044c \u043e\u0446\u0435\u043d\u043a\u0443")



    logger.info(

        "1C rating saved: ext_id=%s chat_id=%s dialog_id=%s appeal_id=%s target=%s rating=%s",

        external_chat_id,

        chat_id,

        dialog_id,

        appeal_id,

        body.target,

        body.rating,

    )

    survey_result = None
    if body.target == "operator":
        survey_result = survey_service.maybe_start_survey_after_employee_csat(dialog_id, appeal_id)
        if not survey_result or not survey_result.get("started"):
            legacy_survey_result = survey_service.maybe_start_survey_after_appeal_closed(dialog_id, appeal_id)
            if legacy_survey_result and legacy_survey_result.get("started"):
                survey_result = legacy_survey_result

    return {

        "status": "ok",

        "dialog_id": dialog_id,

        "appeal_id": appeal_id,

        "target": body.target,

        "rating": body.rating,
        "survey": survey_result,

    }





@router.post("/dialogs/{dialog_id}/favorite")

def mark_dialog_favorite(

    dialog_id: int,

    current_user: Dict[str, object] = Depends(get_current_user),

):

    dialog = database.get_chat_dialog(dialog_id)

    if dialog is None:

        raise HTTPException(status_code=404, detail="\u0414\u0438\u0430\u043b\u043e\u0433 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d")

    if not database.user_can_access_chat(

        current_user["id"], current_user["role"], dialog["chat_id"], dialog_id

    ):

        raise HTTPException(status_code=403, detail="\u041d\u0435\u0442 \u0434\u043e\u0441\u0442\u0443\u043f\u0430 \u043a \u0434\u0438\u0430\u043b\u043e\u0433\u0443")

    database.set_favorite_dialog(current_user["id"], dialog_id, True)

    return {"status": "ok"}





@router.delete("/dialogs/{dialog_id}/favorite")

def unmark_dialog_favorite(

    dialog_id: int,

    current_user: Dict[str, object] = Depends(get_current_user),

):

    dialog = database.get_chat_dialog(dialog_id)

    if dialog is None:

        raise HTTPException(status_code=404, detail="\u0414\u0438\u0430\u043b\u043e\u0433 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d")

    if not database.user_can_access_chat(

        current_user["id"], current_user["role"], dialog["chat_id"], dialog_id

    ):

        raise HTTPException(status_code=403, detail="\u041d\u0435\u0442 \u0434\u043e\u0441\u0442\u0443\u043f\u0430 \u043a \u0434\u0438\u0430\u043b\u043e\u0433\u0443")

    database.set_favorite_dialog(current_user["id"], dialog_id, False)

    return {"status": "ok"}





@router.delete("/chats/{chat_id}")

def delete_chat_endpoint(

    chat_id: int,

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    chat = database.get_chat(chat_id)

    if chat is None:

        raise HTTPException(status_code=404, detail="\u0414\u0438\u0430\u043b\u043e\u0433 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d")

    try:

        database.delete_chat(chat_id)

    except ValueError as exc:

        raise HTTPException(status_code=404, detail=str(exc)) from exc

    return {"status": "ok"}





@router.delete("/dialogs/{dialog_id}")

def delete_dialog_endpoint(

    dialog_id: int,

    _: Dict[str, object] = Depends(require_admin_or_moderator),

):

    try:

        database.delete_chat_dialog(dialog_id)

    except ValueError as exc:

        raise HTTPException(status_code=404, detail=str(exc)) from exc

    return {"status": "ok"}





@router.get("/updates", response_model=List[NotificationResponse])

def list_updates(

    since: Optional[str] = None,

    current_user: Dict[str, object] = Depends(get_current_user),

):

    since_dt: Optional[datetime] = None

    if since:

        try:

            since_dt = datetime.fromisoformat(since)

        except ValueError as exc:

            raise HTTPException(status_code=400, detail="\u041d\u0435\u043a\u043e\u0440\u0440\u0435\u043a\u0442\u043d\u044b\u0439 \u0444\u043e\u0440\u043c\u0430\u0442 \u0432\u0440\u0435\u043c\u0435\u043d\u0438") from exc

    updates = database.list_updates_since(current_user["id"], current_user["role"], since_dt)

    enriched: List[NotificationResponse] = []

    for entry in updates:

        created_at_value = entry.get("created_at")

        if isinstance(created_at_value, datetime):

            created_at_iso = created_at_value.isoformat()

        else:

            created_at_iso = str(created_at_value)

        entry_type = entry.get("type", "message")

        if entry_type == "message":

            section_id = entry.get("section")

            section_title = None

            if section_id:

                section = next((s for s in database.SECTIONS if s["id"] == section_id), None)

                if section:

                    section_title = section["title"]

            enriched.append(

                NotificationResponse(

                    type="message",

                    chat_id=entry.get("chat_id"),

                    chat_title=entry.get("chat_title"),

                    text=entry.get("text", ""),

                    created_at=created_at_iso,

                    section=section_id,

                    section_title=section_title,

                    bin=entry.get("bin"),

                    dialog_id=entry.get("dialog_id"),

                )

            )

            continue

        if entry_type == "bin_assigned":

            bin_value = entry.get("bin") or entry.get("metadata", {}).get("bin")

            message = "\u0412\u0430\u043c \u043d\u0430\u0437\u043d\u0430\u0447\u0435\u043d \u043d\u043e\u0432\u044b\u0439 \u0411\u0418\u041d."

            if bin_value:

                message = f"\u0412\u0430\u043c \u043d\u0430\u0437\u043d\u0430\u0447\u0435\u043d \u0411\u0418\u041d {bin_value}."

            enriched.append(

                NotificationResponse(

                    type="bin_assignment",

                    chat_id=None,

                    chat_title=None,

                    text=message,

                    created_at=created_at_iso,

                    section=None,

                    section_title=None,

                    bin=bin_value,

                )

            )

        else:

            enriched.append(

                NotificationResponse(

                    type=str(entry_type),

                    chat_id=entry.get("chat_id"),

                    chat_title=entry.get("chat_title"),

                    text=entry.get("text", ""),

                    created_at=created_at_iso,

                    section=entry.get("section"),

                    section_title=None,

                    bin=entry.get("bin"),

                    dialog_id=entry.get("dialog_id"),

                )

            )

    return enriched


@router.get("/surveys/templates")
def list_survey_templates_endpoint(
    status: str | None = None,
    _: Dict[str, object] = Depends(require_admin_or_moderator),
):
    return database.list_survey_templates(status=status)


@router.post("/surveys/templates")
def create_survey_template_endpoint(
    body: SurveyTemplateRequest,
    current_user: Dict[str, object] = Depends(require_admin_or_moderator),
):
    try:
        return database.create_survey_template(
            title=body.title,
            description=body.description,
            audience=body.audience,
            status=body.status,
            launch_rules=body.launch_rules,
            trigger_type=body.trigger_type,
            periodic_interval=body.periodic_interval,
            scheduled_at=body.scheduled_at,
            is_anonymous=body.is_anonymous,
            questions=body.questions,
            created_by=int(current_user["id"]),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.put("/surveys/templates/{template_id}")
def update_survey_template_endpoint(
    template_id: int,
    body: SurveyTemplateRequest,
    _: Dict[str, object] = Depends(require_admin_or_moderator),
):
    try:
        return database.update_survey_template(
            int(template_id),
            title=body.title,
            description=body.description,
            audience=body.audience,
            status=body.status,
            launch_rules=body.launch_rules,
            trigger_type=body.trigger_type,
            periodic_interval=body.periodic_interval,
            scheduled_at=body.scheduled_at,
            is_anonymous=body.is_anonymous,
            questions=body.questions,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/surveys/templates/{template_id}/duplicate")
def duplicate_survey_template_endpoint(
    template_id: int,
    current_user: Dict[str, object] = Depends(require_admin_or_moderator),
):
    try:
        return database.duplicate_survey_template(int(template_id), created_by=int(current_user["id"]))
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@router.delete("/surveys/templates/{template_id}")
def delete_survey_template_endpoint(
    template_id: int,
    _: Dict[str, object] = Depends(require_admin_or_moderator),
):
    deleted = database.delete_survey_template(int(template_id))
    if not deleted:
        raise HTTPException(status_code=409, detail="Шаблон нельзя удалить")
    return {"status": "ok"}


@router.post("/surveys/manual-launch")
def launch_survey_endpoint(
    body: ManualSurveyLaunchRequest,
    _: Dict[str, object] = Depends(require_admin_or_moderator),
):
    targets = database.resolve_survey_manual_targets(
        bin_values=body.bin_values,
        dialog_ids=body.dialog_ids,
        only_closed=False,
    )
    started: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    for target in targets:
        session = database.start_survey_session(
            template_id=int(body.template_id),
            chat_id=int(target["chat_id"]),
            dialog_id=int(target["dialog_id"]) if target.get("dialog_id") is not None else None,
            appeal_id=int(target["appeal_id"]) if target.get("appeal_id") is not None else None,
            trigger_source="admin_manual",
        )
        if session:
            started.append(
                {
                    "session_id": int(session["id"]),
                    "chat_id": int(session["chat_id"]),
                    "dialog_id": session.get("dialog_id"),
                    "appeal_id": session.get("appeal_id"),
                    "bin": session.get("bin"),
                }
            )
        else:
            skipped.append(
                {
                    "chat_id": int(target["chat_id"]),
                    "dialog_id": target.get("dialog_id"),
                    "appeal_id": target.get("appeal_id"),
                    "bin": target.get("bin"),
                }
            )
    return {
        "started": started,
        "skipped": skipped,
        "started_count": len(started),
        "skipped_count": len(skipped),
    }


@router.get("/analytics/surveys")
def get_survey_analytics_endpoint(
    audience: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    operator_name: str | None = None,
    bin: str | None = None,
    region: str | None = None,
    topic: str | None = None,
    template_id: int | None = None,
    section: str | None = None,
    _: Dict[str, object] = Depends(require_admin_or_moderator),
):
    return database.get_survey_analytics(
        audience=audience,
        start_date=_parse_optional_date_param(start_date, field_name="start_date"),
        end_date=_parse_optional_date_param(end_date, field_name="end_date"),
        operator_name=operator_name,
        bin_value=bin,
        region=region,
        topic=topic,
        template_id=template_id,
        section=section,
    )


@router.get("/analytics/ratings/summary")
def get_ratings_summary_endpoint(
    _: Dict[str, object] = Depends(require_admin_or_moderator),
):
    return database.get_ratings_summary()


@router.get("/analytics/ratings/employees")
def get_employee_ratings_analytics_endpoint(
    employee_id: int | None = None,
    employee_name: str | None = None,
    _: Dict[str, object] = Depends(require_admin_or_moderator),
):
    return database.get_employee_ratings_analytics_filtered(
        employee_id=employee_id,
        employee_name=employee_name,
    )


@router.get("/analytics/ratings/clients")
def get_client_ratings_analytics_endpoint(
    _: Dict[str, object] = Depends(require_admin_or_moderator),
):
    return database.get_client_ratings_analytics()


@router.get("/analytics/ratings/ai")
def get_ai_ratings_analytics_endpoint(
    _: Dict[str, object] = Depends(require_admin_or_moderator),
):
    return database.get_ai_ratings_analytics()


@router.get("/analytics/ratings/matrix")
def get_mutual_rating_matrix_endpoint(
    _: Dict[str, object] = Depends(require_admin_or_moderator),
):
    return database.get_mutual_rating_matrix()


@router.get("/analytics/ratings/ledger")
def get_rating_ledger_endpoint(
    start_date: str | None = None,
    end_date: str | None = None,
    rater_type: str | None = None,
    rated_object_type: str | None = None,
    employee_id: int | None = None,
    employee_name: str | None = None,
    client_bin: str | None = None,
    client_id: int | None = None,
    section: str | None = None,
    region: str | None = None,
    organization: str | None = None,
    ai_involved: bool | None = None,
    channel: str | None = None,
    limit: int = Query(default=50, ge=1),
    offset: int = Query(default=0, ge=0),
    _: Dict[str, object] = Depends(require_admin_or_moderator),
):
    return database.get_rating_ledger(
        start_date=_parse_optional_date_param(start_date, field_name="start_date"),
        end_date=_parse_optional_date_param(end_date, field_name="end_date"),
        rater_type=rater_type,
        rated_object_type=rated_object_type,
        employee_id=employee_id,
        employee_name=employee_name,
        client_bin=client_bin,
        client_id=client_id,
        section=section,
        region=region,
        organization=organization,
        ai_involved=ai_involved,
        channel=channel,
        limit=limit,
        offset=offset,
    )


@router.get("/analytics/employee-client-assessments")
def get_employee_client_assessment_analytics_endpoint(
    employee_id: int | None = None,
    employee_name: str | None = None,
    client_bin: str | None = None,
    _: Dict[str, object] = Depends(require_admin_or_moderator),
):
    return employee_client_assessments.get_employee_assessment_analytics(
        employee_id=employee_id,
        employee_name=employee_name,
        client_bin=client_bin,
    )


@router.post("/employee-client-assessments/{assessment_id}/submit")
def submit_employee_client_assessment_endpoint(
    assessment_id: int,
    body: EmployeeClientAssessmentSubmitRequest,
    _: Dict[str, object] = Depends(get_current_user),
):
    try:
        return employee_client_assessments.submit_employee_assessment(
            int(assessment_id),
            question_clarity_score=body.question_clarity_score,
            data_completeness_score=body.data_completeness_score,
            client_response_speed_score=body.client_response_speed_score,
            business_communication_score=body.business_communication_score,
            client_readiness_score=body.client_readiness_score,
            low_score_reason=body.low_score_reason,
            internal_comment=body.internal_comment,
            interaction_status=body.interaction_status,
            interaction_flag=body.interaction_flag,
            request_repeat_status=body.request_repeat_status,
            client_data_overdue=body.client_data_overdue,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/health")

def healthcheck() -> Dict[str, str]:

    return {"status": "ok"}





app.include_router(router)



app.include_router(kabinet_backend.router, prefix="/kabinet")




















